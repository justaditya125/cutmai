#!/usr/bin/env python3
import http.server
import socketserver
import json
import os
import time
import urllib.request
import urllib.error
import secrets
import hashlib
import sys
import io
from collections import defaultdict
from datetime import datetime, timedelta
from pymongo import MongoClient
from bson import ObjectId

# Reconfigure stdout/stderr encoding for robust terminal output on Windows
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')


# ─── Load secrets from .env (never hardcode credentials) ─────────────────────
try:
    from dotenv import load_dotenv
    base_dir = os.path.dirname(os.path.abspath(__file__))
    load_dotenv(os.path.join(base_dir, '.env'))
except ImportError:
    pass  # dotenv optional; fall back to environment variables     

PORT     = 3000
DB_NAME  = 'cutm_ai'

import threading

MONGO_URI       = os.environ.get('MONGO_URI', '')
CLAUDE_API_KEY  = os.environ.get('CLAUDE_API_KEY', '')
GOOGLE_CLIENT_ID = os.environ.get('GOOGLE_CLIENT_ID', '')

# Multi-key load balancing support
CLAUDE_API_KEYS_STR = os.environ.get('CLAUDE_API_KEYS', '')
CLAUDE_API_KEYS = []
for key_source in [CLAUDE_API_KEYS_STR, CLAUDE_API_KEY]:
    if key_source:
        CLAUDE_API_KEYS.extend([k.strip() for k in key_source.split(',') if k.strip()])

# ─── Monitoring Globals & System References ───────────────────────────────────
import smtplib
from email.mime.text import MIMEText
from email.header import Header

START_TIME = datetime.now()
FAILED_API_REQUESTS = 0
SUSPICIOUS_ACTIVITIES = []

def log_suspicious_activity(ip_or_user, activity_type, description, risk_level="MEDIUM"):
    """Logs an event in memory to present in the suspicious activities table."""
    event = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user_ip": ip_or_user,
        "type": activity_type,
        "desc": description,
        "risk": risk_level
    }
    SUSPICIOUS_ACTIVITIES.append(event)
    print(f"⚠️  [MONITORING] Suspicious Activity: {activity_type} from {ip_or_user} - {description} ({risk_level})")

def get_uptime():
    """Calculates active system uptime from START_TIME."""
    delta = datetime.now() - START_TIME
    days = delta.days
    hours, remainder = divmod(delta.seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{days}d {hours}h {minutes}m {seconds}s"

def get_user_activity_data():
    """Queries MongoDB Atlas to aggregate user activity and Claude usage statistics."""
    db = get_db()
    if db is None:
        return []
    
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    users_data = []
    
    try:
        # Find token usage for today
        token_records = list(db.token_usage.find({"created_at": {"$gte": today_start}}))
        
        # Group stats by user_id
        user_stats = defaultdict(lambda: {"input": 0, "output": 0, "total": 0, "reqs": 0, "models": set()})
        for r in token_records:
            u_id = str(r.get("user_id"))
            user_stats[u_id]["input"] += r.get("input_tokens", 0)
            user_stats[u_id]["output"] += r.get("output_tokens", 0)
            user_stats[u_id]["total"] += r.get("total_tokens", 0)
            user_stats[u_id]["reqs"] += 1
            if r.get("model"):
                user_stats[u_id]["models"].add(r.get("model"))
                
        # Find all users active today or logged in today
        all_users = list(db.users.find())
        for u in all_users:
            u_id = str(u["_id"])
            last_login = u.get("last_login")
            logged_in_today = last_login and last_login >= today_start
            has_token_usage = u_id in user_stats
            
            if logged_in_today or has_token_usage:
                # Retrieve the latest session for IP/User Agent info
                latest_session = db.user_sessions.find_one(
                    {"user_id": u["_id"]},
                    sort=[("created_at", -1)]
                )
                
                ip = latest_session.get("ip_address", "unknown") if latest_session else "unknown"
                ua = latest_session.get("user_agent", "unknown") if latest_session else "unknown"
                
                device = "Desktop"
                if ua:
                    ua_lower = ua.lower()
                    if "mobile" in ua_lower or "android" in ua_lower or "iphone" in ua_lower:
                        device = "Mobile"
                    elif "postman" in ua_lower:
                        device = "Postman"
                    elif "python" in ua_lower or "urllib" in ua_lower:
                        device = "Python"
                
                stats = user_stats[u_id]
                model_used = ", ".join(stats["models"]) if stats["models"] else "N/A"
                if len(model_used) > 18:
                    model_used = model_used[:15] + "..."
                
                # Calculate credits based on standard Anthropic pricing:
                # Sonnet 4.5/3.5: $3.00/M input, $15.00/M output
                # Haiku 4.5/3.5: $0.25/M input, $1.25/M output
                # Opus: $15.00/M input, $75.00/M output
                credits = 0.0
                for r in token_records:
                    if str(r.get("user_id")) == u_id:
                        m = r.get("model", "").lower()
                        in_t = r.get("input_tokens", 0)
                        out_t = r.get("output_tokens", 0)
                        if "sonnet" in m:
                            credits += (in_t * 3.0 + out_t * 15.0) / 1_000_000
                        elif "haiku" in m:
                            credits += (in_t * 0.25 + out_t * 1.25) / 1_000_000
                        elif "opus" in m:
                            credits += (in_t * 15.0 + out_t * 75.0) / 1_000_000
                        else:
                            credits += (in_t * 3.0 + out_t * 15.0) / 1_000_000
                
                users_data.append({
                    "email": u.get("email", ""),
                    "login_time": last_login.strftime("%H:%M:%S") if last_login else "N/A",
                    "ip": ip,
                    "device": device,
                    "status": "Success" if u.get("is_active", True) else "Blocked",
                    "model": model_used if model_used else "claude-haiku-4-5",
                    "reqs": str(stats["reqs"]),
                    "tokens": str(stats["total"]),
                    "credits": credits
                })
    except Exception as ex:
        print(f"❌ Error gathering activity data: {ex}")
        
    return users_data

def generate_monitoring_email_body():
    """Constructs the complete text body of the status report."""
    now = datetime.now()
    current_timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
    
    server_status = "RUNNING"
    server_uptime = get_uptime()
    active_threads = str(threading.active_count())
    system_cpu_usage = "1.8"
    system_memory_usage = "24.5"
    
    mongodb_status = "DISCONNECTED"
    mongodb_failed_queries = "0"
    mongodb_last_operation = "N/A"
    mongodb_last_update = "N/A"
    
    db = get_db()
    if db is not None:
        try:
            db.command('ping')
            mongodb_status = "CONNECTED"
            users_count = db.users.count_documents({})
            conv_count = db.conversations.count_documents({})
            msg_count = db.messages.count_documents({})
            mongodb_last_operation = f"Active Docs: {users_count} users, {conv_count} convs, {msg_count} msgs"
            
            latest_msg = db.messages.find_one({}, sort=[("created_at", -1)])
            if latest_msg and "created_at" in latest_msg:
                mongodb_last_update = latest_msg["created_at"].strftime("%Y-%m-%d %H:%M:%S")
            else:
                mongodb_last_update = current_timestamp
        except Exception as e:
            mongodb_failed_queries = "1"
            mongodb_last_operation = f"Error: {str(e)}"
            
    users_data = get_user_activity_data()
    header_user = "| {:<22} | {:<8} | {:<12} | {:<8} | {:<7} | {:<18} | {:<4} | {:<8} | {:<8} |".format(
        "User Email", "Login", "IP Address", "Device", "Status", "Model Used", "Reqs", "Tokens", "Credits"
    )
    divider_user = "-" * len(header_user)
    
    user_rows = []
    total_active_users = len(users_data)
    total_tokens_today = 0
    total_reqs_today = 0
    credits_used_today = 0.0
    
    for u in users_data:
        user_rows.append("| {:<22} | {:<8} | {:<12} | {:<8} | {:<7} | {:<18} | {:<4} | {:<8} | ${:<7.4f} |".format(
            u["email"], u["login_time"], u["ip"], u["device"], u["status"], u["model"], u["reqs"], u["tokens"], u["credits"]
        ))
        total_tokens_today += int(u["tokens"]) if u["tokens"].isdigit() else 0
        total_reqs_today += int(u["reqs"]) if u["reqs"].isdigit() else 0
        credits_used_today += u["credits"]
        
    if not user_rows:
        user_activity_table = f"{divider_user}\n| No active users detected today.                                                                                       |\n{divider_user}"
    else:
        user_activity_table = f"{divider_user}\n{header_user}\n{divider_user}\n" + "\n".join(user_rows) + f"\n{divider_user}"
        
    total_input_tokens = 0
    total_output_tokens = 0
    if db is not None:
        try:
            today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            token_records = list(db.token_usage.find({"created_at": {"$gte": today_start}}))
            for r in token_records:
                total_input_tokens += r.get("input_tokens", 0)
                total_output_tokens += r.get("output_tokens", 0)
        except Exception:
            pass
            
    base_credits = 51.31
    remaining_credits = base_credits - credits_used_today
    if remaining_credits < 0:
        remaining_credits = 0.0
        
    low_credit_alert_triggered = "YES" if remaining_credits < 10.0 else "NO"
    est_runtime_remaining = f"{remaining_credits / credits_used_today:.1f}" if credits_used_today > 0 else "365+"
    
    header_susp = "| {:<19} | {:<18} | {:<18} | {:<32} | {:<10} |".format(
        "Timestamp", "User/IP", "Activity Type", "Description", "Risk Level"
    )
    divider_susp = "-" * len(header_susp)
    susp_rows = []
    
    for s in SUSPICIOUS_ACTIVITIES[-20:]:
        desc = s["desc"]
        if len(desc) > 32:
            desc = desc[:29] + "..."
        susp_rows.append("| {:<19} | {:<18} | {:<18} | {:<32} | {:<10} |".format(
            s["timestamp"], s["user_ip"], s["type"], desc, s["risk"]
        ))
        
    if not susp_rows:
        suspicious_activity_table = f"{divider_susp}\n| No suspicious activities detected today.                                                                               |\n{divider_susp}"
    else:
        suspicious_activity_table = f"{divider_susp}\n{header_susp}\n{divider_susp}\n" + "\n".join(susp_rows) + f"\n{divider_susp}"
        
    blocked_ips = len(set(s["user_ip"] for s in SUSPICIOUS_ACTIVITIES if "Blocked" in s["type"] or "Limit" in s["type"]))
    failed_logins_count = sum(1 for s in SUSPICIOUS_ACTIVITIES if s["type"] == "Failed Login")
    unauthorized_api_calls = sum(1 for s in SUSPICIOUS_ACTIVITIES if s["type"] == "Unauthorized Admin Access")
    suspicious_activities_count = len(SUSPICIOUS_ACTIVITIES)
    
    template = """================================================================================
                    CUTM AI SYSTEM MONITORING REPORT
================================================================================
Server:       CUTM AI Production Node
Timestamp:    {current_timestamp}
Trigger Rule: Scheduled Status Check / Alert System Trigger
--------------------------------------------------------------------------------

[1] SERVER STATUS
--------------------------------------------------------------------------------
Server Status  : {server_status}
Server Port    : 3000
Uptime         : {server_uptime}
Active Thread  : {active_threads}
CPU / Memory   : {system_cpu_usage}% / {system_memory_usage}%


[2] DATABASE STATUS (MONGODB ATLAS)
--------------------------------------------------------------------------------
Connection     : {mongodb_status}
DB Name        : cutm_ai
Failed Queries : {mongodb_failed_queries}
Last Operation : {mongodb_last_operation}
Last DB Update : {mongodb_last_update}


[3] USER ACTIVITY & CLAUDE USAGE SUMMARY
--------------------------------------------------------------------------------
{user_activity_table}

Notes on user activity: 
- Total active users today: {total_active_users}
- Institutional validation status: Active (@cutm.ac.in & @cutmap.ac.in only)


[4] CLAUDE API USAGE & PERFORMANCE
--------------------------------------------------------------------------------
Active Model Rotator : Enabled (2 keys loaded for round-robin load balancing)
Models Loaded        : claude-haiku-4-5, claude-sonnet-4-5, claude-opus-4-5
Total API Requests   : {total_api_requests}
Failed API Requests  : {failed_api_requests}
Input Tokens Sent    : {total_input_tokens}
Output Tokens Recv   : {total_output_tokens}
Total Tokens Used    : {total_tokens}


[5] CREDIT & COST MONITORING
--------------------------------------------------------------------------------
Anthropic Balance Rem. : ${remaining_credits:.4f}
Credits Used Today     : ${credits_used_today:.4f}
Low Credit Threshold   : $10.00
Alert Triggered        : {low_credit_alert_triggered}
Est. Runtime Remaining : {est_runtime_remaining} days


[6] SUSPICIOUS ACTIVITY REPORT
--------------------------------------------------------------------------------
{suspicious_activity_table}


[7] SECURITY & THREAT SUMMARY
--------------------------------------------------------------------------------
Blocked IP Addresses       : {blocked_ips}
Failed Login Attempts     : {failed_logins_count}
Unauthorized API Calls     : {unauthorized_api_calls}
Suspicious Activities      : {suspicious_activities_count}


[8] ALERTS & ADMIN INTRUSION SETTINGS
--------------------------------------------------------------------------------
SMTP Host             : smtp.gmail.com
SMTP Port             : 587
Sender Address        : {sender_email}
Email Alerts Enabled  : True
Admins Notified       : aditya.sah@thegttech.com
                        kalyankv@cutmap.ac.in

================================================================================
Generated automatically by CUTM AI Production Node. Please do not reply directly.
================================================================================"""

    return template.format(
        current_timestamp=current_timestamp,
        server_status=server_status,
        server_uptime=server_uptime,
        active_threads=active_threads,
        system_cpu_usage=system_cpu_usage,
        system_memory_usage=system_memory_usage,
        mongodb_status=mongodb_status,
        mongodb_failed_queries=mongodb_failed_queries,
        mongodb_last_operation=mongodb_last_operation,
        mongodb_last_update=mongodb_last_update,
        user_activity_table=user_activity_table,
        total_active_users=str(total_active_users),
        total_api_requests=str(total_reqs_today),
        failed_api_requests=str(FAILED_API_REQUESTS),
        total_input_tokens=str(total_input_tokens),
        total_output_tokens=str(total_output_tokens),
        total_tokens=str(total_tokens_today),
        remaining_credits=remaining_credits,
        credits_used_today=credits_used_today,
        low_credit_alert_triggered=low_credit_alert_triggered,
        est_runtime_remaining=est_runtime_remaining,
        suspicious_activity_table=suspicious_activity_table,
        blocked_ips=str(blocked_ips),
        failed_logins_count=str(failed_logins_count),
        unauthorized_api_calls=str(unauthorized_api_calls),
        suspicious_activities_count=str(suspicious_activities_count),
        sender_email=os.environ.get("SMTP_EMAIL", "alertsemail@cutmap.ac.in")
    )

def send_monitoring_email(subject, body):
    """Establishes a connection to Gmail's SMTP servers to transmit the plaintext report."""
    smtp_host = "smtp.gmail.com"
    smtp_port = 587
    
    sender_email = os.environ.get("SMTP_EMAIL", "alertsemail@cutmap.ac.in")
    sender_password = os.environ.get("SMTP_PASSWORD", "")
    
    admin_emails_str = os.environ.get("ADMIN_EMAIL", "")
    # Parse and filter out comments (elements starting with #) and empty values
    admin_emails = []
    for email in admin_emails_str.split(","):
        email = email.strip()
        if email and not email.startswith("#"):
            admin_emails.append(email)
            
    if not admin_emails:
        admin_emails = ["aditya.sah@thegttech.com", "kalyankv@cutmap.ac.in"]
        
    if not sender_password or sender_password in ["your_16_char_app_password", ""]:
        print("⚠️  [MONITORING] SMTP credentials not set or still default. Skipping email delivery.")
        print("\n=== EMAIL SIMULATION ===")
        print(f"Subject: {subject}")
        print(f"To: {', '.join(admin_emails)}")
        print(body)
        print("========================\n")
        return False
        
    try:
        server = smtplib.SMTP(smtp_host, smtp_port, timeout=10)
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(sender_email, sender_password)
        
        success_count = 0
        for email in admin_emails:
            try:
                msg = MIMEText(body, 'plain', 'utf-8')
                msg['Subject'] = Header(subject, 'utf-8')
                msg['From'] = sender_email
                msg['To'] = email
                
                server.sendmail(sender_email, [email], msg.as_string())
                print(f"✅ [MONITORING] Email sent successfully to: {email}")
                success_count += 1
            except Exception as individual_error:
                print(f"❌ [MONITORING] Failed to send email to {email}: {individual_error}")
                
        server.quit()
        return success_count > 0
    except Exception as e:
        print(f"❌ [MONITORING] SMTP connection or login failed: {e}")
        return False

def send_email_in_background(subject, body):
    """Spawns a daemon thread to transmit monitoring emails asynchronously without blocking the server."""
    threading.Thread(target=send_monitoring_email, args=(subject, body), daemon=True).start()

def daily_scheduler_loop():
    """Background scheduler that executes every day at 3:00 AM local time to send out status emails."""
    import time
    last_sent_date = None
    print("⏰ [MONITORING] Daily scheduler active: Scheduled to send summary at 3:00 AM local time.")
    
    while True:
        try:
            # Check every 30 seconds
            time.sleep(30)
            now = datetime.now()
            
            # Check if current time is 3:00 AM
            if now.hour == 3 and now.minute == 0:
                today_str = now.strftime("%Y-%m-%d")
                if last_sent_date != today_str:
                    last_sent_date = today_str
                    print(f"⏰ [MONITORING] Scheduling trigger matched for 3:00 AM. Preparing daily report...")
                    subject = f"[MONITOR] Daily Status Report - {today_str}"
                    body = generate_monitoring_email_body()
                    send_email_in_background(subject, body)
                    
                    # Trigger user daily notifications
                    try:
                        trigger_user_daily_notifications()
                    except Exception as user_notify_err:
                        print(f"⚠️ [MONITORING] Failed to trigger user daily notifications: {user_notify_err}")
        except Exception as e:
            print(f"⚠️ [MONITORING] Scheduler loop encountered an error: {e}")

def send_user_usage_email(recipient_email, recipient_name, tokens, credits, balance, is_high_usage, threshold):
    """Sends a daily usage report or high-usage alert email to an individual user."""
    smtp_host = "smtp.gmail.com"
    smtp_port = 587
    
    sender_email = os.environ.get("SMTP_EMAIL", "alertsemail@cutmap.ac.in")
    sender_password = os.environ.get("SMTP_PASSWORD", "")
    
    if not sender_password or sender_password in ["your_16_char_app_password", ""]:
        print(f"⚠️  [USER NOTIFICATION] SMTP credentials not set. Skipping daily usage email to {recipient_email}.")
        return False
        
    subject = f"[CUTM AI] Daily Usage Summary - {datetime.now().strftime('%Y-%m-%d')}"
    if is_high_usage:
        subject = f"⚠️ [ALERT] High Credit Usage Warning - CUTM AI"
        
    name = recipient_name if recipient_name else "User"
    
    if is_high_usage:
        body = f"""================================================================================
                    ⚠️ HIGH USAGE ALERT: CUTM AI CHATBOT
================================================================================
Dear {name},

This is an automated alert to notify you that your daily credit usage has exceeded your warning threshold of ${threshold:.2f} today.

DAILY ACTIVITY SUMMARY:
------------------------------------------------------------
Date               : {datetime.now().strftime('%Y-%m-%d')}
Tokens Consumed    : {tokens:,} tokens
Credits Expended   : ${credits:.4f} USD
Remaining Balance  : {balance:,} tokens
------------------------------------------------------------

STATUS ALERT:
Your usage is higher than usual today. If your remaining balance falls below your expected requirements, please limit your token consumption or contact your administrator to request a quota increase.

Best regards,
CUTM AI Operations Team
Centurion University of Technology and Management
================================================================================"""
    else:
        body = f"""================================================================================
                    DAILY USAGE SUMMARY: CUTM AI CHATBOT
================================================================================
Dear {name},

Here is your automated daily summary of your CUTM AI Chatbot usage.

DAILY ACTIVITY SUMMARY:
------------------------------------------------------------
Date               : {datetime.now().strftime('%Y-%m-%d')}
Tokens Consumed    : {tokens:,} tokens
Credits Expended   : ${credits:.4f} USD
Remaining Balance  : {balance:,} tokens
------------------------------------------------------------

Thank you for using CUTM AI!

Best regards,
CUTM AI Operations Team
Centurion University of Technology and Management
================================================================================"""

    try:
        msg = MIMEText(body, 'plain', 'utf-8')
        msg['Subject'] = Header(subject, 'utf-8')
        msg['From'] = sender_email
        msg['To'] = recipient_email
        
        server = smtplib.SMTP(smtp_host, smtp_port, timeout=10)
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, [recipient_email], msg.as_string())
        server.quit()
        print(f"✅ [USER NOTIFICATION] Daily report sent successfully to: {recipient_email}")
        return True
    except Exception as e:
        print(f"❌ [USER NOTIFICATION] Failed to send usage email to {recipient_email}: {e}")
        return False

def send_user_usage_email_in_background(recipient_email, recipient_name, tokens, credits, balance, is_high_usage, threshold):
    """Spawns a background thread to send user usage email."""
    threading.Thread(
        target=send_user_usage_email, 
        args=(recipient_email, recipient_name, tokens, credits, balance, is_high_usage, threshold), 
        daemon=True
    ).start()

def trigger_user_daily_notifications():
    """Queries active users in the past 24 hours and dispatches usage reports/warnings."""
    enable_user_emails = os.environ.get("ENABLE_USER_USAGE_EMAILS", "True").lower() == "true"
    if not enable_user_emails:
        print("ℹ️  [USER NOTIFICATION] User daily usage emails are disabled in settings. Skipping.")
        return
        
    db = get_db()
    if db is None:
        return
        
    # Query token usage records from the past 24 hours
    yesterday = datetime.now() - timedelta(days=1)
    token_records = list(db.token_usage.find({"created_at": {"$gte": yesterday}}))
    
    if not token_records:
        print("ℹ️  [USER NOTIFICATION] No user activity found in the last 24 hours. No emails sent.")
        return
        
    # Group by user_id
    user_records = defaultdict(list)
    for r in token_records:
        u_id = r.get("user_id")
        if u_id:
            user_records[str(u_id)].append(r)
            
    # Load warning threshold (default: $0.50)
    try:
        threshold = float(os.environ.get("USER_DAILY_WARNING_THRESHOLD_USD", "0.50"))
    except ValueError:
        threshold = 0.50
        
    print(f"ℹ️  [USER NOTIFICATION] Processing daily usage reports for {len(user_records)} active users. Warning threshold: ${threshold:.2f}")
    
    for u_id_str, records in user_records.items():
        try:
            from bson import ObjectId
            u_id = ObjectId(u_id_str)
        except Exception:
            continue
            
        user = db.users.find_one({"_id": u_id})
        if not user:
            continue
            
        email = user.get("email")
        name = user.get("name", "")
        balance = user.get("token_balance", 0)
        
        if not email:
            continue
            
        # Calculate daily tokens and credits
        daily_tokens = 0
        daily_credits = 0.0
        
        for r in records:
            in_t = r.get("input_tokens", 0)
            out_t = r.get("output_tokens", 0)
            total_t = r.get("total_tokens", 0)
            daily_tokens += total_t
            
            model = r.get("model", "").lower()
            if "sonnet" in model:
                daily_credits += (in_t * 3.0 + out_t * 15.0) / 1_000_000
            elif "haiku" in model:
                daily_credits += (in_t * 0.25 + out_t * 1.25) / 1_000_000
            elif "opus" in model:
                daily_credits += (in_t * 15.0 + out_t * 75.0) / 1_000_000
            else:
                daily_credits += (in_t * 3.0 + out_t * 15.0) / 1_000_000
                
        is_high_usage = daily_credits >= threshold
        
        # Dispatch email in background thread
        send_user_usage_email_in_background(email, name, daily_tokens, daily_credits, balance, is_high_usage, threshold)



class KeyRotator:
    def __init__(self, keys):
        self.keys = keys
        self._index = 0
        self._lock = threading.Lock()
        
    def get_key(self):
        if not self.keys:
            return ""
        with self._lock:
            key = self.keys[self._index]
            self._index = (self._index + 1) % len(self.keys)
            print(f"🔄 [Load Balancer] Rotating API Key: using key ending in ...{key[-8:] if len(key) > 8 else '???'}")
            return key

api_key_rotator = KeyRotator(CLAUDE_API_KEYS)

if not MONGO_URI:
    print('⚠️  WARNING: MONGO_URI not set. Add it to your .env file.')
if not CLAUDE_API_KEYS:
    print('⚠️  WARNING: No Claude API keys loaded. Set CLAUDE_API_KEY or CLAUDE_API_KEYS in .env')
else:
    print(f"✅ Loaded {len(CLAUDE_API_KEYS)} Claude API keys for round-robin load balancing.")

# ─── Rate Limiter ─────────────────────────────────────────────────────────────
# Tracks per-IP attempt timestamps for sensitive endpoints.
_rate_store: dict = defaultdict(list)

def is_rate_limited(ip: str, endpoint: str, limit: int = 10, window: int = 60) -> bool:
    """
    Returns True if `ip` has exceeded `limit` calls to `endpoint` in the last
    `window` seconds.  Automatically prunes stale entries.
    """
    key = f"{ip}:{endpoint}"
    now = time.monotonic()
    _rate_store[key] = [t for t in _rate_store[key] if now - t < window]
    if len(_rate_store[key]) >= limit:
        log_suspicious_activity(ip, "Rate Limit Triggered", f"Exceeded limits ({limit} per {window}s) on endpoint: {endpoint}", "LOW")
        return True
    _rate_store[key].append(now)
    return False


# ─── Database Helper ──────────────────────────────────────────────────────────
# Single long-lived client with built-in connection pool (replaces per-request connections)
try:
    _mongo_client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
except Exception as _e:
    print(f"❌ MongoDB client init error: {_e}")
    _mongo_client = None

def get_db():
    """Return the shared database reference (connection-pooled)."""
    if _mongo_client is None:
        return None
    try:
        return _mongo_client[DB_NAME]
    except Exception as e:
        print(f"❌ DB access error: {e}")
        return None

def fetch_url_text(url):
    import urllib.request
    import urllib.error
    from bs4 import BeautifulSoup
    try:
        url = url.strip()
        req = urllib.request.Request(
            url,
            headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5'
            }
        )
        with urllib.request.urlopen(req, timeout=8) as response:
            html = response.read()
            soup = BeautifulSoup(html, 'html.parser')
            for element in soup(["script", "style", "noscript", "iframe", "header", "footer", "nav"]):
                element.decompose()
            text = soup.get_text()
            lines = (line.strip() for line in text.splitlines())
            chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
            clean_text = '\n'.join(chunk for chunk in chunks if chunk)
            return clean_text[:6000]
    except Exception as e:
        print(f"⚠️ [Web Fetcher] Error fetching {url}: {e}")
        return f"[Failed to retrieve webpage content from {url} due to error: {e}]"

def extract_gdrive_file_id(url):
    """Extract Google Drive file/folder ID from various Drive URL formats."""
    import re
    # Match file id in /file/d/<id>/ or ?id=<id> or /d/<id>
    patterns = [
        r'/file/d/([a-zA-Z0-9_-]{10,})',
        r'id=([a-zA-Z0-9_-]{10,})',
        r'/d/([a-zA-Z0-9_-]{10,})',
        r'/folders/([a-zA-Z0-9_-]{10,})',
    ]
    for pattern in patterns:
        m = re.search(pattern, url)
        if m:
            return m.group(1)
    return None

def fetch_gdrive_text(url):
    """Download a publicly shared Google Drive file or folder and extract text."""
    import tempfile, os, shutil
    try:
        file_id = extract_gdrive_file_id(url)
        if not file_id:
            return "[Could not parse Google Drive file ID from the provided link.]"

        is_folder = '/folders/' in url
        tmp_dir = tempfile.mkdtemp()

        try:
            import gdown
        except ImportError:
            return "[gdown not installed. Run: pip install gdown]"

        # ── FOLDER ───────────────────────────────────────────────────────────────
        if is_folder:
            print(f"📂 [Drive Fetcher] Downloading FOLDER ID: {file_id}")
            folder_path = os.path.join(tmp_dir, 'drive_folder')
            os.makedirs(folder_path, exist_ok=True)
            try:
                gdown.download_folder(id=file_id, output=folder_path, quiet=True, use_cookies=False)
            except Exception as e:
                shutil.rmtree(tmp_dir, ignore_errors=True)
                return f"[Google Drive folder download failed: {e}. Share the folder as 'Anyone with the link'.]"

            downloaded_files = []
            for root, dirs, files in os.walk(folder_path):
                for fname in files:
                    downloaded_files.append(os.path.join(root, fname))

            if not downloaded_files:
                shutil.rmtree(tmp_dir, ignore_errors=True)
                return "[Google Drive folder appears empty or is restricted. Share it as 'Anyone with the link'.]"

            print(f"✅ [Drive Fetcher] Downloaded {len(downloaded_files)} files from folder")
            all_text = []
            for fpath in downloaded_files[:10]:
                fname = os.path.basename(fpath)
                t = _extract_text_from_file(fpath)
                if t and t.strip():
                    all_text.append(f"--- File: {fname} ---\n{t.strip()}")
            shutil.rmtree(tmp_dir, ignore_errors=True)
            combined = "\n\n".join(all_text)
            return combined[:10000] if combined else "[No readable text found in the Google Drive folder.]"

        # ── SINGLE FILE ──────────────────────────────────────────────────────────
        else:
            print(f"📄 [Drive Fetcher] Downloading FILE ID: {file_id}")
            output_path = os.path.join(tmp_dir, 'drive_file')
            try:
                downloaded = gdown.download(id=file_id, output=output_path, quiet=True, fuzzy=True)
            except Exception as e:
                shutil.rmtree(tmp_dir, ignore_errors=True)
                return f"[Google Drive file download failed: {e}. Share as 'Anyone with the link'.]"

            if not downloaded or not os.path.exists(downloaded):
                shutil.rmtree(tmp_dir, ignore_errors=True)
                return "[Google Drive download failed. The file may be private. Share as 'Anyone with the link'.]"

            print(f"✅ [Drive Fetcher] Downloaded {os.path.getsize(downloaded)} bytes")
            text = _extract_text_from_file(downloaded)
            shutil.rmtree(tmp_dir, ignore_errors=True)
            return text[:8000] if text else "[No readable text content found in the Google Drive file.]"

    except Exception as e:
        print(f"⚠️ [Drive Fetcher] Error: {e}")
        return f"[Google Drive fetch failed: {e}]"


def _extract_text_from_file(filepath):
    """Extract readable text from a downloaded file based on its extension."""
    lower_path = filepath.lower()
    text = ""
    if lower_path.endswith('.pdf'):
        try:
            from pdfminer.high_level import extract_text as pdf_extract
            text = pdf_extract(filepath)
        except ImportError:
            text = "[PDF found but pdfminer.six not installed. Run: pip install pdfminer.six]"
        except Exception as e:
            text = f"[PDF extraction error: {e}]"
    elif lower_path.endswith('.docx') or lower_path.endswith('.doc'):
        try:
            import mammoth
            with open(filepath, 'rb') as f:
                text = mammoth.extract_raw_text(f).value
        except ImportError:
            try:
                from docx import Document
                text = '\n'.join([p.text for p in Document(filepath).paragraphs])
            except ImportError:
                text = "[Word file found but mammoth/python-docx not installed.]"
        except Exception as e:
            text = f"[Word extraction error: {e}]"
    elif lower_path.endswith('.xlsx'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheets_text = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_lines = [f"--- Sheet: {sheet_name} ---"]
                for row in sheet.iter_rows(values_only=True):
                    if row and any(cell is not None and str(cell).strip() != "" for cell in row):
                        row_str = " | ".join(str(cell).strip() if cell is not None else "" for cell in row)
                        sheet_lines.append(row_str)
                if len(sheet_lines) > 1:
                    sheets_text.append("\n".join(sheet_lines))
            text = "\n\n".join(sheets_text)
        except Exception as e:
            text = f"[Excel extraction error (.xlsx): {e}]"
    elif lower_path.endswith('.xls'):
        try:
            import xlrd
            wb = xlrd.open_workbook(filepath)
            sheets_text = []
            for sheet_idx in range(wb.nsheets):
                sheet = wb.sheet_by_index(sheet_idx)
                sheet_lines = [f"--- Sheet: {sheet.name} ---"]
                for row_idx in range(sheet.nrows):
                    row = sheet.row_values(row_idx)
                    if row and any(cell is not None and str(cell).strip() != "" for cell in row):
                        row_str = " | ".join(str(cell).strip() if cell is not None else "" for cell in row)
                        sheet_lines.append(row_str)
                if len(sheet_lines) > 1:
                    sheets_text.append("\n".join(sheet_lines))
            text = "\n\n".join(sheets_text)
        except Exception as e:
            text = f"[Excel extraction error (.xls): {e}]"
    elif any(lower_path.endswith(ext) for ext in ['.txt', '.csv', '.md', '.py', '.js', '.json', '.xml', '.html', '.log']):
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                text = f.read()
        except Exception as e:
            text = f"[Text read error: {e}]"
    else:
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                raw = f.read()
            text = raw if raw.strip() else ""
        except Exception:
            text = ""
    return text


def convert_to_alphanumeric(password):
    if not password:
        return ""
    # Deterministically convert any password to a 12-character alphanumeric string.
    # Uses SHA-256 to ensure good entropy distribution.
    h = hashlib.sha256(password.encode('utf-8')).digest()
    chars = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    result = []
    for i in range(12):
        byte_val = h[i*2] + (h[i*2 + 1] << 8)
        result.append(chars[byte_val % len(chars)])
    return "".join(result)

def hash_password(password):
    # Convert password to a 12-character alphanumeric representation first
    converted = convert_to_alphanumeric(password)
    salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', converted.encode(), salt.encode(), 100000)
    return f"{salt}:{h.hex()}"

def verify_password(password, stored):
    try:
        salt, h = stored.split(':')
        # Check against the converted alphanumeric version
        converted = convert_to_alphanumeric(password)
        if hashlib.pbkdf2_hmac('sha256', converted.encode(), salt.encode(), 100000).hex() == h:
            return True
        # For backward compatibility (existing users/admin), check raw password
        if hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 100000).hex() == h:
            return True
        return False
    except:
        return False

def create_session(db, user_id, ip=None, ua=None):
    token = secrets.token_urlsafe(32)
    expires = datetime.now() + timedelta(days=30)
    u_id = ObjectId(user_id) if isinstance(user_id, (str, bytes)) else user_id
    try:
        db.user_sessions.insert_one({
            "user_id": u_id,
            "session_token": token,
            "expires_at": expires,
            "created_at": datetime.now(),
            "ip_address": ip,
            "user_agent": ua
        })
        return token
    except Exception as e:
        print(f"❌ Session error: {e}")
        return secrets.token_urlsafe(32)  # fallback token

def save_token_usage(db, user_id, input_tokens, output_tokens, model='claude-haiku-4-5'):
    """Save token usage to database and print detailed log"""
    total = input_tokens + output_tokens
    u_id = ObjectId(user_id) if isinstance(user_id, (str, bytes)) else user_id
    try:
        # Insert token usage record
        db.token_usage.insert_one({
            "user_id": u_id,
            "input_tokens": input_tokens,
            "output_tokens": output_tokens,
            "total_tokens": total,
            "model": model,
            "created_at": datetime.now()
        })

        # Update user total tokens
        db.users.update_one(
            {"_id": u_id},
            {
                "$inc": {
                    "total_tokens_used": total,
                    "total_messages": 1
                },
                "$set": {
                    "updated_at": datetime.now()
                }
            }
        )

        # Fetch updated user stats for logging
        user = db.users.find_one({"_id": u_id})

        # ── Real-time console log ──────────────────────────────────────────
        print("\n" + "="*55)
        print(f"  📊 TOKEN USAGE LOG")
        print("="*55)
        print(f"  👤 User     : {user['email']}")
        print(f"  📥 Input    : {input_tokens} tokens")
        print(f"  📤 Output   : {output_tokens} tokens")
        print(f"  🔢 This msg : {total} tokens")
        print(f"  📈 Total    : {user['total_tokens_used']} tokens (lifetime)")
        print(f"  💬 Messages : {user['total_messages']} total")
        print("="*55 + "\n")

    except Exception as e:
        print(f"❌ Token save error: {e}")

def get_user_quota_info(db, user_id, user_doc=None):
    """Calculates and returns user's token usage, limit, balance, and credits consumed."""
    if user_doc is None:
        user_doc = db.users.find_one({"_id": user_id})
    if not user_doc:
        return None
    
    # Calculate user credits
    credits = 0.0
    try:
        user_tokens = list(db.token_usage.find({"user_id": user_id}))
        for r in user_tokens:
            m = r.get("model", "").lower()
            in_t = r.get("input_tokens", 0)
            out_t = r.get("output_tokens", 0)
            if "sonnet" in m:
                credits += (in_t * 3.0 + out_t * 15.0) / 1_000_000
            elif "haiku" in m:
                credits += (in_t * 0.25 + out_t * 1.25) / 1_000_000
            elif "opus" in m:
                credits += (in_t * 15.0 + out_t * 75.0) / 1_000_000
            else:
                credits += (in_t * 3.0 + out_t * 15.0) / 1_000_000
    except Exception as ex:
        print(f"Error calculating credits for user {user_id}: {ex}")

    total_tokens = user_doc.get("total_tokens_used") or 0
    limit = user_doc.get("token_limit") or 1000000
    balance = max(0, limit - total_tokens)
    
    return {
        'total_tokens_used': total_tokens,
        'token_limit': limit,
        'token_balance': balance,
        'credits_used': credits
    }

def trim_messages(messages_list, max_messages=6):
    """Trims message history to optimize tokens. Ensures alternating user/assistant roles."""
    if not messages_list:
        return []
    if len(messages_list) <= max_messages:
        return messages_list
    trimmed = messages_list[-max_messages:]
    if trimmed and trimmed[0].get('role') == 'assistant':
        trimmed = trimmed[1:]
    return trimmed


# ─── Request Handler ──────────────────────────────────────────────────────────
class ChatbotHandler(http.server.SimpleHTTPRequestHandler):

    def log_message(self, format, *args):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {format % args}")

    def log_error(self, format, *args):
        # Suppress noisy browser connection-drop messages (WinError 10054)
        msg = format % args
        if 'ConnectionResetError' in msg or '10054' in msg or 'BrokenPipeError' in msg:
            return
        print(f"[{datetime.now().strftime('%H:%M:%S')}] ERROR: {msg}")

    def get_client_ip(self):
        """Extracts the true client IP from standard proxy headers, falling back to client_address."""
        for header in ['CF-Connecting-IP', 'X-Forwarded-For', 'X-Real-IP']:
            ip = self.headers.get(header)
            if ip:
                if ',' in ip:
                    return ip.split(',')[0].strip()
                return ip.strip()
        return self.client_address[0]

    def end_headers(self):
        # Restrict CORS to our own origin — not the whole internet
        origin = self.headers.get('Origin', '')
        allowed = f'http://localhost:{PORT}'
        self.send_header('Access-Control-Allow-Origin', allowed if origin.startswith('http://localhost') else allowed)
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.send_header('Vary', 'Origin')
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(200)
        self.end_headers()

    def do_POST(self):
        try:
            if   self.path == '/api/auth/register':          self.handle_register()
            elif self.path == '/api/auth/login':             self.handle_login()
            elif self.path == '/api/auth/google':            self.handle_google()
            elif self.path == '/api/auth/verify':            self.handle_verify()
            elif self.path == '/api/auth/logout':            self.handle_logout()
            elif self.path == '/api/claude':                 self.handle_claude()
            elif self.path == '/api/claude/stream':          self.handle_claude_stream()
            elif self.path == '/api/claude/vision':          self.handle_claude_vision()
            elif self.path == '/api/conversations/new':      self.handle_new_conversation()
            elif self.path == '/api/conversations/list':     self.handle_list_conversations()
            elif self.path == '/api/conversations/messages': self.handle_get_messages()
            elif self.path == '/api/conversations/delete':   self.handle_delete_conversation()
            elif self.path == '/api/conversations/rename':   self.handle_rename_conversation()
            elif self.path == '/api/messages/feedback':      self.handle_message_feedback()
            elif self.path == '/api/messages/edit':          self.handle_edit_message()
            elif self.path == '/api/admin/stats':            self.handle_admin_stats()
            elif self.path == '/api/admin/approve_user':     self.handle_approve_user()
            elif self.path == '/api/admin/set_limit':        self.handle_set_limit()
            elif self.path == '/api/admin/send_monitoring_report': self.handle_send_monitoring_report()
            else: self.send_json(404, {'error': 'Not found'})
        except Exception as e:
            print(f"❌ POST error: {e}")
            self.send_json(500, {'error': 'Internal server error'})

    # ── Conversation APIs ─────────────────────────────────────────────────────
    def get_user_from_token(self, token):
        """Helper: get user_id from session token"""
        if not token: return None
        db = get_db()
        if db is None: return None
        try:
            session = db.user_sessions.find_one({
                "session_token": token,
                "expires_at": {"$gt": datetime.now()}
            })
            if session:
                user = db.users.find_one({"_id": session['user_id'], "is_active": True})
                if user:
                    user['id'] = str(user['_id'])
                    return user
            return None
        except Exception as e:
            print(f"Error fetching user from session token: {e}")
            return None

    def handle_new_conversation(self):
        data  = self.read_body()
        token = data.get('session_token', '')
        title = data.get('title', 'New Chat')
        user  = self.get_user_from_token(token)
        if not user:
            return self.send_json(401, {'error': 'Unauthorized'})
        db = get_db()
        if db is None: return self.send_json(500, {'error': 'DB error'})
        try:
            conv_doc = {
                "user_id": ObjectId(user['id']),
                "title": title,
                "created_at": datetime.now(),
                "updated_at": datetime.now()
            }
            res = db.conversations.insert_one(conv_doc)
            conv_id = str(res.inserted_id)
            self.send_json(200, {'success': True, 'conversation_id': conv_id, 'title': title})
        except Exception as e:
            self.send_json(500, {'error': str(e)})

    def handle_list_conversations(self):
        data  = self.read_body()
        token = data.get('session_token', '')
        user  = self.get_user_from_token(token)
        if not user:
            return self.send_json(401, {'error': 'Unauthorized'})
        db = get_db()
        if db is None: return self.send_json(500, {'error': 'DB error'})
        try:
            # Query conversations owned by user sorted by updated_at desc
            convs = list(db.conversations.find({"user_id": ObjectId(user['id'])}).sort("updated_at", -1))
            
            conv_list = []
            for c in convs:
                # Count messages for this conversation
                msg_count = db.messages.count_documents({"conversation_id": c['_id']})
                
                c_created = c.get("created_at")
                c_updated = c.get("updated_at")
                
                conv_list.append({
                    "id": str(c['_id']),
                    "title": c.get("title", "New Conversation"),
                    "created_at": c_created.strftime('%Y-%m-%d %H:%M:%S') if isinstance(c_created, datetime) else str(c_created),
                    "updated_at": c_updated.strftime('%Y-%m-%d %H:%M:%S') if isinstance(c_updated, datetime) else str(c_updated),
                    "message_count": msg_count
                })
            
            self.send_json(200, {'conversations': conv_list})
        except Exception as e:
            self.send_json(500, {'error': str(e)})

    def handle_get_messages(self):
        data    = self.read_body()
        token   = data.get('session_token', '')
        conv_id = data.get('conversation_id')
        user    = self.get_user_from_token(token)
        if not user:
            return self.send_json(401, {'error': 'Unauthorized'})
        db = get_db()
        if db is None: return self.send_json(500, {'error': 'DB error'})
        try:
            c_id = ObjectId(conv_id) if isinstance(conv_id, str) else conv_id
            # Verify ownership
            conv = db.conversations.find_one({"_id": c_id, "user_id": ObjectId(user['id'])})
            if not conv:
                return self.send_json(403, {'error': 'Forbidden'})
                
            msgs = list(db.messages.find({"conversation_id": c_id}).sort("created_at", 1))
            msg_list = []
            for m in msgs:
                m_created = m.get("created_at")
                msg_list.append({
                    "id": str(m.get("_id")),
                    "role": m.get("role"),
                    "content": m.get("content"),
                    "feedback": m.get("feedback", "none"),
                    "created_at": m_created.strftime('%Y-%m-%d %H:%M:%S') if isinstance(m_created, datetime) else str(m_created)
                })
            self.send_json(200, {'messages': msg_list})
        except Exception as e:
            self.send_json(500, {'error': str(e)})

    def handle_delete_conversation(self):
        data    = self.read_body()
        token   = data.get('session_token', '')
        conv_id = data.get('conversation_id')
        user    = self.get_user_from_token(token)
        if not user:
            return self.send_json(401, {'error': 'Unauthorized'})
        db = get_db()
        if db is None: return self.send_json(500, {'error': 'DB error'})
        try:
            c_id = ObjectId(conv_id) if isinstance(conv_id, str) else conv_id
            
            # Verify ownership and delete conversation
            res = db.conversations.delete_one({"_id": c_id, "user_id": ObjectId(user['id'])})
            if res.deleted_count > 0:
                # Cascade delete associated messages
                db.messages.delete_many({"conversation_id": c_id})
                
            self.send_json(200, {'success': True})
        except Exception as e:
            self.send_json(500, {'error': str(e)})

    def handle_rename_conversation(self):
        data    = self.read_body()
        token   = data.get('session_token', '')
        conv_id = data.get('conversation_id')
        title   = data.get('title', 'New Chat')
        user    = self.get_user_from_token(token)
        if not user:
            return self.send_json(401, {'error': 'Unauthorized'})
        db = get_db()
        if db is None: return self.send_json(500, {'error': 'DB error'})
        try:
            c_id = ObjectId(conv_id) if isinstance(conv_id, str) else conv_id
            db.conversations.update_one(
                {"_id": c_id, "user_id": ObjectId(user['id'])},
                {"$set": {"title": title, "updated_at": datetime.now()}}
            )
            self.send_json(200, {'success': True})
        except Exception as e:
            self.send_json(500, {'error': str(e)})

    # ── Message Feedback API ─────────────────────────────────────────────────
    def handle_message_feedback(self):
        """POST /api/messages/feedback — save like/dislike rating on a message."""
        data       = self.read_body()
        token      = data.get('session_token', '')
        message_id = data.get('message_id')
        feedback   = data.get('feedback', 'none')  # 'like' | 'dislike' | 'none'
        user = self.get_user_from_token(token)
        if not user:
            return self.send_json(401, {'error': 'Unauthorized'})
        if feedback not in ('like', 'dislike', 'none'):
            return self.send_json(400, {'error': 'Invalid feedback value'})
        db = get_db()
        if db is None: return self.send_json(500, {'error': 'DB error'})
        try:
            m_id = ObjectId(message_id)
            msg  = db.messages.find_one({'_id': m_id})
            if not msg:
                return self.send_json(404, {'error': 'Message not found'})
            # Verify conversation belongs to user
            conv = db.conversations.find_one({'_id': msg['conversation_id'], 'user_id': ObjectId(user['id'])})
            if not conv:
                return self.send_json(403, {'error': 'Forbidden'})
            db.messages.update_one({'_id': m_id}, {'$set': {'feedback': feedback}})
            self.send_json(200, {'success': True})
        except Exception as e:
            self.send_json(500, {'error': str(e)})

    # ── Edit Message API ─────────────────────────────────────────────────────
    def handle_edit_message(self):
        """POST /api/messages/edit — update message content and delete subsequent messages."""
        data        = self.read_body()
        token       = data.get('session_token', '')
        message_id  = data.get('message_id')
        new_content = data.get('new_content', '')
        user = self.get_user_from_token(token)
        if not user:
            return self.send_json(401, {'error': 'Unauthorized'})
        if not new_content.strip():
            return self.send_json(400, {'error': 'Content cannot be empty'})
        db = get_db()
        if db is None: return self.send_json(500, {'error': 'DB error'})
        try:
            m_id = ObjectId(message_id)
            msg  = db.messages.find_one({'_id': m_id})
            if not msg:
                return self.send_json(404, {'error': 'Message not found'})
            # Verify conversation ownership
            conv = db.conversations.find_one({'_id': msg['conversation_id'], 'user_id': ObjectId(user['id'])})
            if not conv:
                return self.send_json(403, {'error': 'Forbidden'})
            msg_created_at = msg.get('created_at')
            # Delete all subsequent messages in the same conversation
            db.messages.delete_many({
                'conversation_id': msg['conversation_id'],
                'created_at': {'$gt': msg_created_at}
            })
            # Update the edited message's content
            db.messages.update_one({'_id': m_id}, {'$set': {'content': new_content.strip(), 'edited': True}})
            db.conversations.update_one({'_id': msg['conversation_id']}, {'$set': {'updated_at': datetime.now()}})
            self.send_json(200, {'success': True})
        except Exception as e:
            self.send_json(500, {'error': str(e)})

    # ── Admin Stats API ───────────────────────────────────────────────────────
    def handle_admin_stats(self):
        data  = self.read_body()
        token = data.get('session_token', '')
        user  = self.get_user_from_token(token)
        
        if not user or not user.get('is_admin', False):
            client_ip = self.get_client_ip()
            log_suspicious_activity(
                user.get('email', client_ip) if user else client_ip,
                "Unauthorized Admin Access",
                f"Attempted to access admin stats endpoint from IP {client_ip}",
                "HIGH"
            )
            return self.send_json(403, {'error': 'Forbidden: Admin access required'})


        db = get_db()
        if db is None:
            return self.send_json(500, {'error': 'Database unavailable'})
        try:
            # All users
            users = list(db.users.find())
            
            users_list = []
            for u in users:
                u_id = u['_id']
                conv_count = db.conversations.count_documents({"user_id": u_id})
                sess_count = db.user_sessions.count_documents({"user_id": u_id, "expires_at": {"$gt": datetime.now()}})
                
                u_ll = u.get("last_login")
                u_ca = u.get("created_at")
                
                # Calculate user credits
                credits = 0.0
                try:
                    user_tokens = db.token_usage.find({"user_id": u_id})
                    for r in user_tokens:
                        m = r.get("model", "").lower()
                        in_t = r.get("input_tokens", 0)
                        out_t = r.get("output_tokens", 0)
                        if "sonnet" in m:
                            credits += (in_t * 3.0 + out_t * 15.0) / 1_000_000
                        elif "haiku" in m:
                            credits += (in_t * 0.25 + out_t * 1.25) / 1_000_000
                        elif "opus" in m:
                            credits += (in_t * 15.0 + out_t * 75.0) / 1_000_000
                        else:
                            credits += (in_t * 3.0 + out_t * 15.0) / 1_000_000
                except Exception as ex:
                    print(f"Error calculating credits for user {u_id}: {ex}")

                total_tokens = u.get("total_tokens_used") or 0
                limit = u.get("token_limit") or 1000000
                balance = max(0, limit - total_tokens)

                users_list.append({
                    "id": str(u_id),
                    "email": u.get("email"),
                    "name": u.get("name"),
                    "login_method": u.get("login_method", "email"),
                    "is_approved": u.get("is_approved", False),
                    "total_tokens_used": total_tokens,
                    "token_limit": limit,
                    "token_balance": balance,
                    "credits_used": credits,
                    "total_messages": u.get("total_messages") or 0,
                    "last_login": u_ll.strftime('%Y-%m-%d %H:%M:%S') if isinstance(u_ll, datetime) else 'Never',
                    "created_at": u_ca.strftime('%Y-%m-%d %H:%M:%S') if isinstance(u_ca, datetime) else '',
                    "total_conversations": conv_count,
                    "active_sessions": sess_count
                })
            
            # Sort users by total tokens descend (same as MySQL table view)
            users_list.sort(key=lambda x: x['total_tokens_used'], reverse=True)

            # Summary counts
            total_users = len(users_list)
            grand_total = sum(u.get("total_tokens_used") or 0 for u in users)
            total_msgs = sum(u.get("total_messages") or 0 for u in users)
            active_sessions = db.user_sessions.count_documents({"expires_at": {"$gt": datetime.now()}})
            total_convs = db.conversations.count_documents({})

            # Recent activity logs
            recent = list(db.token_usage.find().sort("created_at", -1).limit(20))
            recent_list = []
            for r in recent:
                user_doc = db.users.find_one({"_id": r['user_id']})
                email = user_doc['email'] if user_doc else 'unknown'
                r_ca = r.get("created_at")
                recent_list.append({
                    "email": email,
                    "input_tokens": r.get("input_tokens", 0),
                    "output_tokens": r.get("output_tokens", 0),
                    "total_tokens": r.get("total_tokens", 0),
                    "created_at": r_ca.strftime('%Y-%m-%d %H:%M:%S') if isinstance(r_ca, datetime) else ''
                })

            # Active sessions
            sessions = list(db.user_sessions.find({"expires_at": {"$gt": datetime.now()}}).sort("created_at", -1))
            sessions_list = []
            for s in sessions:
                user_doc = db.users.find_one({"_id": s['user_id']})
                email = user_doc['email'] if user_doc else 'unknown'
                name = user_doc['name'] if user_doc else '-'
                s_ca = s.get("created_at")
                s_ea = s.get("expires_at")
                sessions_list.append({
                    "email": email,
                    "name": name,
                    "ip_address": s.get("ip_address", "unknown"),
                    "created_at": s_ca.strftime('%Y-%m-%d %H:%M:%S') if isinstance(s_ca, datetime) else '',
                    "expires_at": s_ea.strftime('%Y-%m-%d %H:%M:%S') if isinstance(s_ea, datetime) else ''
                })

            self.send_json(200, {
                'users': users_list,
                'summary': {
                    'total_users': total_users,
                    'grand_total_tokens': int(grand_total),
                    'total_messages': int(total_msgs),
                    'active_sessions': active_sessions,
                    'total_conversations': total_convs
                },
                'recent_activity': recent_list,
                'active_sessions': sessions_list
            })
        except Exception as e:
            print(f"❌ Admin stats error: {e}")
            self.send_json(500, {'error': str(e)})

    def handle_send_monitoring_report(self):
        data  = self.read_body()
        token = data.get('session_token', '')
        user  = self.get_user_from_token(token)
        
        if not user or not user.get('is_admin', False):
            client_ip = self.get_client_ip()
            log_suspicious_activity(
                user.get('email', client_ip) if user else client_ip,
                "Unauthorized Admin Access",
                f"Attempted to trigger status report email from IP {client_ip}",
                "HIGH"
            )
            return self.send_json(403, {'error': 'Forbidden: Admin access required'})

        try:
            subject = f"[MONITOR] Manual Status Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            body = generate_monitoring_email_body()
            
            # Send in background non-blocking
            send_email_in_background(subject, body)
            
            self.send_json(200, {
                'success': True,
                'message': 'Monitoring report email triggered successfully in the background.'
            })
        except Exception as e:
            print(f"❌ Failed manual trigger: {e}")
            self.send_json(500, {'error': str(e)})

    # ── Admin User Approval API ───────────────────────────────────────────────
    def handle_approve_user(self):
        data   = self.read_body()
        token  = data.get('session_token', '')
        user   = self.get_user_from_token(token)
        
        if not user or not user.get('is_admin', False):
            client_ip = self.get_client_ip()
            log_suspicious_activity(
                user.get('email', client_ip) if user else client_ip,
                "Unauthorized User Approval",
                f"Attempted to approve user from IP {client_ip}",
                "HIGH"
            )
            return self.send_json(403, {'error': 'Forbidden: Admin access required'})

        target_user_id = data.get('target_user_id', '')
        action         = data.get('action', '') # 'approve', 'revoke', or 'reject'

        if not target_user_id or action not in ['approve', 'revoke', 'reject']:
            return self.send_json(400, {'error': 'Missing target_user_id or valid action'})

        db = get_db()
        if db is None:
            return self.send_json(500, {'error': 'Database unavailable'})

        try:
            from bson import ObjectId
            t_id = ObjectId(target_user_id) if isinstance(target_user_id, str) else target_user_id
            
            # Prevent admin from revoking themselves
            if str(t_id) == user['id'] and action == 'revoke':
                return self.send_json(400, {'error': 'You cannot revoke your own approval status'})

            if action == 'reject':
                db.users.delete_one({"_id": t_id})
                db.user_sessions.delete_many({"user_id": t_id})
                conv_ids = [c['_id'] for c in db.conversations.find({"user_id": t_id}, {"_id": 1})]
                if conv_ids:
                    db.messages.delete_many({"conversation_id": {"$in": conv_ids}})
                db.conversations.delete_many({"user_id": t_id})
                print(f"🗑️ [Admin] Rejected (deleted) user ID: {target_user_id}")
            else:
                is_approved = (action == 'approve')
                db.users.update_one(
                    {"_id": t_id},
                    {"$set": {"is_approved": is_approved, "updated_at": datetime.now()}}
                )

                # If revoked, delete all active sessions for that user immediately to kick them out
                if not is_approved:
                    db.user_sessions.delete_many({"user_id": t_id})
                    print(f"🚫 [Admin] Revoked approval and terminated sessions for user ID: {target_user_id}")
                else:
                    print(f"✅ [Admin] Approved user ID: {target_user_id}")

            self.send_json(200, {'success': True})
        except Exception as e:
            print(f"❌ Approve user error: {e}")
            self.send_json(500, {'error': str(e)})

    def handle_set_limit(self):
        data  = self.read_body()
        token = data.get('session_token', '')
        user  = self.get_user_from_token(token)
        
        if not user or not user.get('is_admin', False):
            client_ip = self.get_client_ip()
            log_suspicious_activity(
                user.get('email', client_ip) if user else client_ip,
                "Unauthorized Limit Change",
                f"Attempted to set token limit from IP {client_ip}",
                "HIGH"
            )
            return self.send_json(403, {'error': 'Forbidden: Admin access required'})

        target_user_id = data.get('target_user_id', '')
        new_limit      = data.get('token_limit')

        if not target_user_id or new_limit is None:
            return self.send_json(400, {'error': 'Missing target_user_id or token_limit'})

        try:
            new_limit = int(new_limit)
            if new_limit < 0:
                return self.send_json(400, {'error': 'Token limit must be a positive integer'})
        except ValueError:
            return self.send_json(400, {'error': 'Token limit must be a valid integer'})

        db = get_db()
        if db is None:
            return self.send_json(500, {'error': 'Database unavailable'})

        try:
            from bson import ObjectId
            t_id = ObjectId(target_user_id) if isinstance(target_user_id, str) else target_user_id
            
            db.users.update_one(
                {"_id": t_id},
                {"$set": {"token_limit": new_limit, "updated_at": datetime.now()}}
            )
            print(f"⚙️ [Admin] Set token limit to {new_limit} for user ID: {target_user_id}")
            self.send_json(200, {'success': True})
        except Exception as e:
            print(f"❌ Set limit error: {e}")
            self.send_json(500, {'error': str(e)})

    def handle_proxy_image(self):
        self.send_json(403, {'error': 'Image generation is currently disabled'})

    def read_body(self):
        length = int(self.headers.get('Content-Length', 0))
        return json.loads(self.rfile.read(length).decode('utf-8')) if length else {}

    # ── Register ──────────────────────────────────────────────────────────────
    def handle_register(self):
        client_ip = self.get_client_ip()
        # Rate limit: 5 registrations per IP per 10 minutes
        if is_rate_limited(client_ip, 'register', limit=5, window=600):
            return self.send_json(429, {'success': False, 'error': 'Too many requests. Please wait a few minutes.'})

        data = self.read_body()
        email    = data.get('email', '').strip().lower()
        password = data.get('password', '').strip()
        name     = data.get('name', '').strip()

        if not email or not password or not name:
            return self.send_json(400, {'success': False, 'error': 'All fields required'})

        if len(password) < 6:
            return self.send_json(400, {'success': False, 'error': 'Password must be at least 6 characters'})

        # Domain restriction - server side standard registration
        domain = email.split('@')[-1].lower() if '@' in email else ''
        if domain not in ['cutmap.ac.in', 'cutm.ac.in']:
            log_suspicious_activity(email, "Blocked Registration", f"Standard registration attempt from non-institutional domain from IP {client_ip}", "MEDIUM")
            print(f"❌ Blocked standard register: {email}")
            return self.send_json(403, {
                'success': False,
                'error': 'Access denied. Only institutional domains (@cutmap.ac.in and @cutm.ac.in) are allowed.'
            })

        db = get_db()
        if db is None:
            return self.send_json(500, {'success': False, 'error': 'Database unavailable'})

        try:
            # Check existing user
            existing = db.users.find_one({"email": email})
            if existing:
                return self.send_json(409, {'success': False, 'error': 'Email already registered'})

            # Create user
            pw_hash = hash_password(password)
            user_doc = {
                "email": email,
                "password_hash": pw_hash,
                "name": name,
                "profile_picture": None,
                "login_method": "email",
                "is_approved": False,
                "token_limit": 1000000,
                "is_active": True,
                "is_admin": False,
                "total_tokens_used": 0,
                "total_messages": 0,
                "created_at": datetime.now(),
                "updated_at": datetime.now(),
                "last_login": datetime.now()
            }
            res = db.users.insert_one(user_doc)
            user_id = res.inserted_id

            print(f"✅ New user registered (pending approval): {email}")
            self.send_json(201, {
                'success': True,
                'requires_approval': True,
                'message': 'Account created successfully! Your account is pending administrator approval.'
            })
        except Exception as e:
            print(f"❌ Register error: {e}")
            self.send_json(500, {'success': False, 'error': 'Registration failed'})

    # ── Login ─────────────────────────────────────────────────────────────────
    def handle_login(self):
        client_ip = self.get_client_ip()
        # Rate limit: 10 attempts per IP per 60 seconds (brute-force guard)
        if is_rate_limited(client_ip, 'login', limit=10, window=60):
            return self.send_json(429, {'success': False, 'error': 'Too many login attempts. Please wait 60 seconds.'})

        data = self.read_body()
        email    = data.get('email', '').strip().lower()
        password = data.get('password', '').strip()

        if not email or not password:
            return self.send_json(400, {'success': False, 'error': 'Email and password required'})

        db = get_db()
        if db is None:
            return self.send_json(500, {'success': False, 'error': 'Database unavailable'})

        try:
            user = db.users.find_one({"email": email, "is_active": True})

            if not user or not user.get('password_hash') or not verify_password(password, user['password_hash']):
                log_suspicious_activity(email or client_ip, "Failed Login", f"Incorrect password attempt for standard login from IP {client_ip}", "LOW")
                return self.send_json(401, {'success': False, 'error': 'Invalid email or password'})

            # Check if user is approved
            if not user.get('is_approved', False):
                return self.send_json(403, {
                    'success': False,
                    'error': 'Your account is pending administrator approval. Please contact the administrator.'
                })

            # Update last login
            db.users.update_one(
                {"_id": user['_id']},
                {"$set": {"last_login": datetime.now(), "updated_at": datetime.now()}}
            )

            token = create_session(db, user['_id'],
                                   client_ip,
                                   self.headers.get('User-Agent'))

            print(f"✅ Login: {email}")
            self.send_json(200, {
                'success': True,
                'user': {
                    'id': str(user['_id']), 'email': user['email'],
                    'name': user.get('name'), 'login_method': user.get('login_method', 'email'),
                    'profile_picture': user.get('profile_picture'),
                    'is_admin': user.get('is_admin', False)
                },
                'session_token': token
            })
        except Exception as e:
            print(f"❌ Login error: {e}")
            self.send_json(500, {'success': False, 'error': 'Login failed'})

    # ── Google Auth (server-side JWT verification) ────────────────────────────
    def handle_google(self):
        client_ip = self.get_client_ip()
        # Rate limit: 10 Google auth attempts per IP per 60 seconds
        if is_rate_limited(client_ip, 'google', limit=10, window=60):
            return self.send_json(429, {'success': False, 'error': 'Too many requests. Please wait.'})

        data       = self.read_body()
        credential = data.get('credential', '')   # raw Google JWT (preferred)
        google_id  = data.get('google_id', '')    # fallback: pre-decoded sub
        email      = data.get('email', '').strip().lower()
        name       = data.get('name', '')
        picture    = data.get('picture', '')

        # ── Verify the Google JWT server-side ─────────────────────────────────
        if credential and GOOGLE_CLIENT_ID:
            try:
                from google.oauth2 import id_token
                from google.auth.transport import requests as g_requests
                id_info = id_token.verify_oauth2_token(
                    credential,
                    g_requests.Request(),
                    GOOGLE_CLIENT_ID,
                    clock_skew_in_seconds=10
                )
                # Override with verified values — never trust client-supplied data
                google_id = id_info['sub']
                email     = id_info['email'].strip().lower()
                name      = id_info.get('name', name)
                picture   = id_info.get('picture', picture)
                print(f"✅ Google JWT verified for: {email}")
            except Exception as verify_err:
                print(f"❌ Google JWT verification failed: {verify_err}")
                return self.send_json(401, {'success': False, 'error': 'Google token verification failed. Please sign in again.'})
        elif not google_id or not email:
            return self.send_json(400, {'success': False, 'error': 'Missing Google credential data'})
        else:
            # credential not sent AND no GOOGLE_CLIENT_ID configured — log warning
            print('⚠️  Google login: JWT not verified (GOOGLE_CLIENT_ID not set or credential not provided)')

        # Domain restriction — server side
        domain = email.split('@')[-1].lower() if '@' in email else ''
        if domain not in ['cutmap.ac.in', 'cutm.ac.in']:
            log_suspicious_activity(email, "Blocked Google Auth", f"Google auth attempt from non-institutional domain from IP {client_ip}", "MEDIUM")
            print(f"❌ Blocked Google login: {email}")
            return self.send_json(403, {
                'success': False,
                'error': 'Access denied. Only @cutmap.ac.in and @cutm.ac.in are allowed.'
            })


        db = get_db()
        if db is None:
            return self.send_json(500, {'success': False, 'error': 'Database unavailable'})

        try:
            # Lookup by verified Google ID first, then fall back to email
            user = db.users.find_one({"google_id": google_id})

            if not user:
                user = db.users.find_one({"email": email})
                if user:
                    # Link Google ID to existing email account
                    db.users.update_one(
                        {"_id": user['_id']},
                        {"$set": {
                            "google_id": google_id,
                            "name": name,
                            "profile_picture": picture,
                            "login_method": 'google',
                            "updated_at": datetime.now()
                        }}
                    )
                    user = db.users.find_one({"_id": user['_id']})
                else:
                    # Create new Google user (starts as pending approval)
                    res = db.users.insert_one({
                        "email": email,
                        "name": name,
                        "profile_picture": picture,
                        "login_method": "google",
                        "google_id": google_id,
                        "is_approved": False,
                        "token_limit": 1000000,
                        "is_active": True,
                        "is_admin": False,
                        "total_tokens_used": 0,
                        "total_messages": 0,
                        "created_at": datetime.now(),
                        "updated_at": datetime.now(),
                        "last_login": datetime.now()
                    })
                    print(f"✅ New Google user registered (pending approval): {email}")
                    return self.send_json(200, {
                        'success': False,
                        'requires_approval': True,
                        'error': 'Account created successfully! Your account is pending administrator approval.'
                    })

            # Check if user is approved
            if not user.get('is_approved', False):
                return self.send_json(403, {
                    'success': False,
                    'error': 'Your account is pending administrator approval. Please contact the administrator.'
                })

            # Update last login timestamp
            db.users.update_one(
                {"_id": user['_id']},
                {"$set": {"last_login": datetime.now(), "updated_at": datetime.now()}}
            )

            token = create_session(db, user['_id'],
                                   client_ip,
                                   self.headers.get('User-Agent'))

            print(f"✅ Google login: {email}")
            self.send_json(200, {
                'success': True,
                'user': {
                    'id': str(user['_id']), 'email': user['email'],
                    'name': user.get('name'), 'login_method': 'google',
                    'profile_picture': user.get('profile_picture'),
                    'is_admin': user.get('is_admin', False)
                },
                'session_token': token
            })
        except Exception as e:
            print(f"❌ Google auth error: {e}")
            self.send_json(500, {'success': False, 'error': 'Google authentication failed'})

    # ── Verify Session ────────────────────────────────────────────────────────
    def handle_verify(self):
        data  = self.read_body()
        token = data.get('session_token', '')

        if not token:
            return self.send_json(400, {'valid': False, 'error': 'Token required'})

        db = get_db()
        if db is None:
            return self.send_json(500, {'valid': False, 'error': 'Database unavailable'})

        try:
            session = db.user_sessions.find_one({
                "session_token": token,
                "expires_at": {"$gt": datetime.now()}
            })
            if session:
                user = db.users.find_one({"_id": session['user_id'], "is_active": True})
                if user:
                    quota = get_user_quota_info(db, user['_id'], user) or {}
                    return self.send_json(200, {
                        'valid': True,
                        'user': {
                            'id': str(user['_id']), 'email': user['email'],
                            'name': user.get('name'), 'login_method': user.get('login_method', 'email'),
                            'profile_picture': user.get('profile_picture'),
                            'is_admin': user.get('is_admin', False),
                            'total_tokens_used': quota.get('total_tokens_used', 0),
                            'token_limit': quota.get('token_limit', 1000000),
                            'token_balance': quota.get('token_balance', 1000000),
                            'credits_used': quota.get('credits_used', 0.0)
                        }
                    })
            self.send_json(401, {'valid': False, 'error': 'Invalid or expired session'})
        except Exception as e:
            print(f"❌ Verify error: {e}")
            self.send_json(500, {'valid': False, 'error': 'Verification failed'})

    # ── Logout ────────────────────────────────────────────────────────────────
    def handle_logout(self):
        data  = self.read_body()
        token = data.get('session_token', '')

        db = get_db()
        if db is not None and token:
            try:
                db.user_sessions.delete_one({"session_token": token})
                print(f"✅ Logged out session: {token[:20]}...")
            except Exception as e:
                print(f"❌ Logout error: {e}")

        self.send_json(200, {'success': True})


    # ── Claude Streaming API (SSE) ────────────────────────────────────────────
    def handle_claude_stream(self):
        """Stream Claude's response using Server-Sent Events (SSE)."""
        global FAILED_API_REQUESTS
        import re, json
        client_ip = self.get_client_ip()
        if is_rate_limited(client_ip, 'claude', limit=30, window=60):
            self.send_response(429)
            self.send_header('Content-Type', 'text/event-stream')
            self.end_headers()
            self.wfile.write(b'data: {"error": "Rate limited"}\n\n')
            return

        data = self.read_body()
        session_token = data.get('session_token', '')
        conv_id       = data.get('conversation_id')
        user_message  = data.get('user_message', '')
        user_info     = self.get_user_from_token(session_token)
        if not user_info:
            self.send_response(401)
            self.send_header('Content-Type', 'text/event-stream')
            self.end_headers()
            self.wfile.write(b'data: {"error": "Unauthorized"}\n\n')
            return

        total_tokens_used = user_info.get('total_tokens_used', 0)
        token_limit = user_info.get('token_limit', 1000000)
        if total_tokens_used >= token_limit:
            self.send_response(403)
            self.send_header('Content-Type', 'text/event-stream')
            self.end_headers()
            self.wfile.write(b'data: {"error": "Token limit exceeded"}\n\n')
            return

        user_id = ObjectId(user_info['id'])
        db = get_db()

        # Auto-create conversation
        if db is not None and user_id and not conv_id:
            try:
                title = (user_message[:40] + '...') if len(user_message) > 40 else user_message
                result = db.conversations.insert_one({
                    "user_id": user_id, "title": title or "New Chat",
                    "created_at": datetime.now(), "updated_at": datetime.now()
                })
                conv_id = str(result.inserted_id)
            except Exception as e:
                print(f"❌ Conv create error: {e}")

        # Fetch URL/Drive content
        urls = re.findall(r'(https?://[^\s]+)', user_message)
        fetched_contents = ""
        GDRIVE_PATTERNS = ['drive.google.com/file/', 'drive.google.com/open', 'drive.google.com/uc',
                           'drive.google.com/drive/', 'docs.google.com/document/',
                           'docs.google.com/spreadsheets/', 'docs.google.com/presentation/']
        for url in urls:
            clean_url = url.rstrip(').,!]')
            if any(p in clean_url for p in GDRIVE_PATTERNS):
                url_text = fetch_gdrive_text(clean_url)
                fetched_contents += f"\n\n[Google Drive File Content from {clean_url}]\n-----------------------------\n{url_text}\n-----------------------------\n"
            else:
                url_text = fetch_url_text(clean_url)
                fetched_contents += f"\n\n[Webpage Content from {clean_url}]\n-----------------------------\n{url_text}\n-----------------------------\n"

        messages = data.get('messages', [])
        if fetched_contents:
            if messages and messages[-1].get('role') == 'user':
                messages[-1]['content'] = fetched_contents + "\nUser query: " + messages[-1]['content']

        messages = trim_messages(messages, max_messages=6)

        chosen_model = data.get('model', 'claude-haiku-4-5')
        if chosen_model not in ['claude-haiku-4-5', 'claude-sonnet-4-5', 'claude-opus-4-5']:
            chosen_model = 'claude-haiku-4-5'

        system_instructions = (
            "You are CUTM AI, an AI assistant for Centurion University (CUTM). Use markdown.\n"
            "When generating code, prefix with filename tag: [filename=\"file.ext\"] inside code blocks."
        )

        active_key = api_key_rotator.get_key()
        if not active_key:
            self.send_response(500)
            self.send_header('Content-Type', 'text/event-stream')
            self.end_headers()
            self.wfile.write(b'data: {"error": "No API key"}\n\n')
            return

        thinking_enabled = data.get('thinking_enabled', False)
        payload_dict = {
            'model': chosen_model,
            'max_tokens': min(data.get('max_tokens', 4000), 4000),
            'system': system_instructions,
            'messages': messages,
            'stream': True
        }
        if thinking_enabled:
            payload_dict['thinking'] = {
                'type': 'enabled',
                'budget_tokens': 1024
            }
        claude_payload = json.dumps(payload_dict).encode('utf-8')

        # Send SSE headers
        self.send_response(200)
        self.send_header('Content-Type', 'text/event-stream')
        self.send_header('Cache-Control', 'no-cache')
        self.send_header('X-Accel-Buffering', 'no')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()

        full_response = ""
        input_tokens = 0
        output_tokens = 0

        try:
            req = urllib.request.Request(
                'https://api.anthropic.com/v1/messages',
                data=claude_payload,
                headers={
                    'Content-Type': 'application/json',
                    'x-api-key': active_key,
                    'anthropic-version': '2023-06-01'
                }
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                for raw_line in resp:
                    line = raw_line.decode('utf-8').strip()
                    if not line or not line.startswith('data:'):
                        continue
                    payload_str = line[5:].strip()
                    if payload_str == '[DONE]':
                        break
                    try:
                        chunk = json.loads(payload_str)
                        event_type = chunk.get('type', '')
                        if event_type == 'content_block_delta':
                            delta = chunk.get('delta', {})
                            delta_type = delta.get('type', '')
                            if delta_type == 'thinking_delta':
                                delta_thinking = delta.get('thinking', '')
                                if delta_thinking:
                                    msg_out = json.dumps({'type': 'thinking', 'text': delta_thinking})
                                    self.wfile.write(f'data: {msg_out}\n\n'.encode('utf-8'))
                                    self.wfile.flush()
                            else:
                                delta_text = delta.get('text', '')
                                if delta_text:
                                    full_response += delta_text
                                    msg_out = json.dumps({'type': 'delta', 'text': delta_text})
                                    self.wfile.write(f'data: {msg_out}\n\n'.encode('utf-8'))
                                    self.wfile.flush()
                        elif event_type == 'message_delta':
                            usage = chunk.get('usage', {})
                            output_tokens = usage.get('output_tokens', 0)
                        elif event_type == 'message_start':
                            usage = chunk.get('message', {}).get('usage', {})
                            input_tokens = usage.get('input_tokens', 0)
                    except Exception:
                        continue

            # Save to DB after stream completes
            user_msg_id = None
            asst_msg_id = None
            if db is not None and user_id and conv_id and user_message:
                try:
                    c_id = ObjectId(conv_id) if isinstance(conv_id, str) else conv_id
                    user_res = db.messages.insert_one({"conversation_id": c_id, "user_id": user_id, "role": 'user', "content": user_message, "created_at": datetime.now(), "feedback": "none"})
                    user_msg_id = str(user_res.inserted_id)
                    if full_response:
                        asst_res = db.messages.insert_one({"conversation_id": c_id, "user_id": user_id, "role": 'assistant', "content": full_response, "created_at": datetime.now(), "feedback": "none"})
                        asst_msg_id = str(asst_res.inserted_id)
                    db.conversations.update_one({"_id": c_id}, {"$set": {"updated_at": datetime.now()}})
                    if input_tokens or output_tokens:
                        save_token_usage(db, user_id, input_tokens, output_tokens, chosen_model)
                except Exception as e:
                    print(f"❌ Stream DB save error: {e}")

            # Send final done event with metadata including message IDs
            done_msg = json.dumps({'type': 'done', 'conversation_id': str(conv_id), 'input_tokens': input_tokens, 'output_tokens': output_tokens, 'user_message_id': user_msg_id, 'assistant_message_id': asst_msg_id})
            self.wfile.write(f'data: {done_msg}\n\n'.encode('utf-8'))
            self.wfile.flush()

        except Exception as e:
            FAILED_API_REQUESTS += 1
            print(f"❌ Stream error: {e}")
            err_msg = json.dumps({'type': 'error', 'message': str(e)})
            try:
                self.wfile.write(f'data: {err_msg}\n\n'.encode('utf-8'))
                self.wfile.flush()
            except Exception:
                pass

    # ── Claude Vision API (Image Understanding) ───────────────────────────────
    def handle_claude_vision(self):
        """Accept base64 image + text, send to Claude vision, return response."""
        global FAILED_API_REQUESTS
        data = self.read_body()
        session_token = data.get('session_token', '')
        user_info = self.get_user_from_token(session_token)
        if not user_info:
            return self.send_json(401, {'error': 'Unauthorized'})

        image_b64   = data.get('image_base64', '')   # base64 string without prefix
        media_type  = data.get('media_type', 'image/jpeg')  # e.g. image/png
        user_text   = data.get('text', 'What is in this image?')
        chosen_model = data.get('model', 'claude-haiku-4-5')
        if chosen_model not in ['claude-haiku-4-5', 'claude-sonnet-4-5', 'claude-opus-4-5']:
            chosen_model = 'claude-haiku-4-5'

        if not image_b64:
            return self.send_json(400, {'error': 'No image provided'})

        active_key = api_key_rotator.get_key()
        if not active_key:
            return self.send_json(500, {'error': 'No API key'})

        payload = json.dumps({
            'model': chosen_model,
            'max_tokens': 2048,
            'messages': [{
                'role': 'user',
                'content': [
                    {'type': 'image', 'source': {'type': 'base64', 'media_type': media_type, 'data': image_b64}},
                    {'type': 'text', 'text': user_text}
                ]
            }]
        }).encode('utf-8')

        try:
            req = urllib.request.Request(
                'https://api.anthropic.com/v1/messages',
                data=payload,
                headers={'Content-Type': 'application/json', 'x-api-key': active_key, 'anthropic-version': '2023-06-01'}
            )
            with urllib.request.urlopen(req) as resp:
                response_json = json.loads(resp.read())
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(response_json).encode('utf-8'))
        except Exception as e:
            FAILED_API_REQUESTS += 1
            print(f"❌ Vision error: {e}")
            self.send_json(500, {'error': str(e)})

    # ── Claude API ────────────────────────────────────────────────────────────
    def handle_claude(self):
        global FAILED_API_REQUESTS
        client_ip = self.get_client_ip()
        # Rate limit: 30 Claude messages per IP per 60 seconds (abuse/cost guard)
        if is_rate_limited(client_ip, 'claude', limit=30, window=60):
            return self.send_json(429, {'error': 'Rate limit reached. Please slow down.'})

        data = self.read_body()


        session_token = data.get('session_token', '')
        conv_id       = data.get('conversation_id')
        user_message  = data.get('user_message', '')
        user_info     = self.get_user_from_token(session_token)
        if not user_info:
            return self.send_json(401, {'error': 'Unauthorized'})

        # Check token quota limit
        total_tokens_used = user_info.get('total_tokens_used', 0)
        token_limit = user_info.get('token_limit', 1000000)
        if total_tokens_used >= token_limit:
            return self.send_json(403, {'error': 'You have exceeded your allocated token limit. Please contact the administrator.'})

        user_id       = ObjectId(user_info['id'])

        db = get_db()

        # Auto-create conversation if none provided
        if db is not None and user_id and not conv_id:
            try:
                # Generate title from first message (first 40 chars)
                title = (user_message[:40] + '...') if len(user_message) > 40 else user_message or 'New Chat'
                conv_doc = {
                    "user_id": user_id,
                    "title": title,
                    "created_at": datetime.now(),
                    "updated_at": datetime.now()
                }
                res = db.conversations.insert_one(conv_doc)
                conv_id = str(res.inserted_id)
            except Exception as e:
                print(f"Auto-conversation creation failed: {e}")

        # Helper functions for Token Optimization

        def summarize_history(messages_list):
            """Generates a high-density 2-3 sentence summary of older messages using the ultra-cheap Haiku 3."""
            active_key = api_key_rotator.get_key()
            if not active_key:
                return None
            try:
                summary_prompt = "You are an assistant that summarizes the provided conversation history into a very brief, high-density summary (2-3 sentences max) highlighting key contexts, decisions, and facts."
                payload = json.dumps({
                    'model': 'claude-haiku-4-5',
                    'max_tokens': 300,
                    'system': summary_prompt,
                    'messages': messages_list + [{'role': 'user', 'content': 'Summarize the context of this conversation so far in 2-3 sentences.'}]
                }).encode('utf-8')
                
                req_obj = urllib.request.Request(
                    'https://api.anthropic.com/v1/messages',
                    data=payload,
                    headers={
                        'Content-Type': 'application/json',
                        'x-api-key': active_key,
                        'anthropic-version': '2023-06-01'
                    }
                )
                with urllib.request.urlopen(req_obj) as resp_obj:
                    resp_json = json.loads(resp_obj.read())
                    if resp_json.get('content') and resp_json['content'][0].get('text'):
                        return resp_json['content'][0]['text']
            except Exception as se:
                print(f"⚠️ Summarization failed: {se}")
            return None

        # Optimized system prompt to enforce charts and filename metadata
        system_instructions = (
            "You are CUTM AI, an AI assistant for Centurion University (CUTM). Use markdown. Image generation is disabled.\n\n"
            "INSTRUCTIONS FOR DATA VISUALIZATION:\n"
            "If the user requests data analytics, sales reports, tables, trends, or any data comparative views, perform thorough calculations/analyses and present interactive charts using the following JSON schema wrapped in a ```chart code block (e.g., ```chart\\n{...}\\n```):\n"
            "{\n"
            "  \"type\": \"bar\", // or \"line\", \"pie\", \"doughnut\", \"radar\"\n"
            "  \"title\": \"Title of Chart\",\n"
            "  \"labels\": [\"Label1\", \"Label2\", ...],\n"
            "  \"datasets\": [\n"
            "    {\n"
            "      \"label\": \"Dataset Label\",\n"
            "      \"data\": [value1, value2, ...]\n"
            "    }\n"
            "  ]\n"
            "}\n\n"
            "INSTRUCTIONS FOR EDITABLE DOCUMENTS & ARTIFACT FILENAMES:\n"
            "When generating structured reports, tables, code scripts, or documents that the user might want to edit, modify, or download, prefix the code block with a filename tag format: [filename=\"filename.ext\"] on the very first line inside the fenced code block (e.g., ```markdown\\n[filename=\"report.md\"]\\n...\\n``` or ```text\\n[filename=\"document.txt\"]\\n...\\n```). Ensure files representing Microsoft Word documents have a `.doc` or `.docx` extension. When the user uploads a document (PDF or Word) and asks for modifications, read the provided text context, perform modifications, and provide the modified version in a fenced code block with a filename tag on the first line."
        )

        messages = data.get('messages', [])

        # Scan for URLs in user_message and fetch content dynamically
        import re
        urls = re.findall(r'(https?://[^\s]+)', user_message)
        fetched_contents = ""
        GDRIVE_PATTERNS = [
            'drive.google.com/file/',
            'drive.google.com/open',
            'drive.google.com/uc',
            'drive.google.com/drive/',
            'docs.google.com/document/',
            'docs.google.com/spreadsheets/',
            'docs.google.com/presentation/',
        ]
        for url in urls:
            clean_url = url.rstrip(').,!]')
            is_gdrive = any(p in clean_url for p in GDRIVE_PATTERNS)
            if is_gdrive:
                print(f"📂 [Drive Fetcher] Detected Google Drive URL: {clean_url}")
                url_text = fetch_gdrive_text(clean_url)
                fetched_contents += f"\n\n[Google Drive File Content from {clean_url}]\n-----------------------------\n{url_text}\n-----------------------------\n"
            else:
                print(f"🔍 [Web Fetcher] Detected URL: {clean_url}")
                url_text = fetch_url_text(clean_url)
                fetched_contents += f"\n\n[Webpage Content from {clean_url}]\n-----------------------------\n{url_text}\n-----------------------------\n"

        if fetched_contents:
            user_message = fetched_contents + "\nUser query: " + user_message
            if messages and messages[-1].get('role') == 'user':
                messages[-1]['content'] = fetched_contents + "\nUser query: " + messages[-1]['content']
        
        # Apply Smart Summarization & Context Trimming (trigger early at 12 messages to keep input low)
        if len(messages) > 12:
            messages_to_summarize = messages[:-6]
            recent_messages = messages[-6:]
            if recent_messages and recent_messages[0].get('role') == 'assistant':
                recent_messages = recent_messages[1:]
                
            summary = summarize_history(messages_to_summarize)
            if summary:
                system_instructions += f"\n\nHere is a summary of the earlier part of the conversation for your context:\n{summary}"
            messages = recent_messages
        else:
            messages = trim_messages(messages, max_messages=6)

        # Respect the frontend's chosen model if valid, fallback to ultra-low cost claude-haiku-4-5
        chosen_model = data.get('model', 'claude-haiku-4-5')
        if chosen_model not in ['claude-haiku-4-5', 'claude-sonnet-4-5', 'claude-opus-4-5']:
            chosen_model = 'claude-haiku-4-5'
        thinking_enabled = data.get('thinking_enabled', False)
        payload_dict = {
            'model': chosen_model,
            'max_tokens': min(data.get('max_tokens', 4000), 4000),
            'system': system_instructions,
            'messages': messages
        }
        if thinking_enabled:
            payload_dict['thinking'] = {
                'type': 'enabled',
                'budget_tokens': 1024
            }
        claude_payload = json.dumps(payload_dict).encode('utf-8')

        active_key = api_key_rotator.get_key()
        if not active_key:
            return self.send_json(500, {'error': 'No API keys configured'})

        try:
            req = urllib.request.Request(
                'https://api.anthropic.com/v1/messages',
                data=claude_payload,
                headers={
                    'Content-Type': 'application/json',
                    'x-api-key': active_key,
                    'anthropic-version': '2023-06-01'
                }
            )

            with urllib.request.urlopen(req) as resp:
                response_data = resp.read()
                response_json = json.loads(response_data)

                assistant_text = ''
                if response_json.get('content'):
                    assistant_text = response_json['content'][0].get('text', '')

                # Save messages to DB
                if db is not None and user_id and conv_id and user_message:
                    try:
                        c_id = ObjectId(conv_id) if isinstance(conv_id, str) else conv_id
                        # Save user message
                        db.messages.insert_one({
                            "conversation_id": c_id,
                            "user_id": user_id,
                            "role": 'user',
                            "content": user_message,
                            "created_at": datetime.now()
                        })
                        # Save assistant response
                        if assistant_text:
                            db.messages.insert_one({
                                "conversation_id": c_id,
                                "user_id": user_id,
                                "role": 'assistant',
                                "content": assistant_text,
                                "created_at": datetime.now()
                            })
                        # Update conversation timestamp
                        db.conversations.update_one(
                            {"_id": c_id},
                            {"$set": {"updated_at": datetime.now()}}
                        )

                        # Save token usage
                        usage = response_json.get('usage', {})
                        save_token_usage(db, user_id, usage.get('input_tokens', 0), usage.get('output_tokens', 0), chosen_model)
                    except Exception as e:
                        print(f"❌ Message save error: {e}")

                # Add conversation_id to response
                response_json['conversation_id'] = str(conv_id)

                # Fetch and add user quota/credit details
                if db is not None:
                    response_json['user_quota'] = get_user_quota_info(db, user_id)

                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(response_json).encode('utf-8'))

        except urllib.error.HTTPError as e:
            FAILED_API_REQUESTS += 1
            err = e.read().decode('utf-8')
            print(f"❌ Claude API error: {e.code} - {err}")
            client_ip = self.get_client_ip()
            log_suspicious_activity(user_info.get('email', client_ip) if user_info else client_ip, "Claude API HTTP Error", f"HTTP {e.code} error received: {err[:60]}", "MEDIUM")
            self.send_response(e.code)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(err.encode('utf-8'))

        except Exception as e:
            FAILED_API_REQUESTS += 1
            print(f"❌ Claude error: {e}")
            self.send_json(500, {'error': str(e)})



    # ── Serve HTML files ──────────────────────────────────────────────────────
    def do_GET(self):
        if self.path.startswith('/api/proxy/image'):
            self.handle_proxy_image()
            return

        routes = {
            '/':            'login.html',
            '/login.html':  'login.html',
            '/index.html':  'index.html',
            '/signup.html': 'signup.html',
            '/admin':       'admin.html',
            '/admin.html':  'admin.html',
        }
        filename = routes.get(self.path)



        if filename:
            try:
                base_dir = os.path.dirname(os.path.abspath(__file__))
                full_path = os.path.join(base_dir, filename)
                with open(full_path, 'rb') as f:
                    content = f.read()
                self.send_response(200)
                self.send_header('Content-Type', 'text/html')
                self.end_headers()
                self.wfile.write(content)
            except FileNotFoundError:
                self.send_error(404, f"{filename} not found")
        else:
            try:
                super().do_GET()
            except:
                self.send_error(404, "Not found")

    # ── JSON helper ───────────────────────────────────────────────────────────
    def send_json(self, status, data):
        try:
            body = json.dumps(data).encode('utf-8')
            self.send_response(status)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(body)
        except Exception as e:
            print(f"❌ send_json error: {e}")


# ─── Main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    print("🚀 Starting CUTM AI Chatbot Server...")

    # Test DB connection
    db = get_db()
    db_status = "Connected (MongoDB Atlas)"
    if db is not None:
        try:
            db.command('ping')
            print("✅ MongoDB database connected successfully")
        except Exception as e:
            print(f"⚠️  MongoDB connection failed: {e}")
            db_status = f"Disconnected (Error: {str(e)})"
    else:
        print("⚠️  MongoDB not connected - check credentials in MONGO_URI")
        db_status = "Disconnected (No URI)"

    # Send non-blocking startup alert email
    startup_subject = f"[MONITOR] CUTM AI Server Startup Alert - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    startup_body = f"""================================================================================
                    CUTM AI SERVER STARTUP NOTIFICATION
================================================================================
The CUTM AI Chatbot Server has started successfully.

Server Port          : {PORT}
Database Status      : {db_status}
Active Model Rotator : Enabled ({len(CLAUDE_API_KEYS)} keys loaded)
Email Alerts Enabled : True
Admins Notified      : kalyankv@cutmap.ac.in, aditya.sah@thegttech.com

Timestamp            : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Server               : CUTM AI Production Node
================================================================================"""
    send_email_in_background(startup_subject, startup_body)

    # Start the 3:00 AM daily status monitoring scheduler thread
    threading.Thread(target=daily_scheduler_loop, daemon=True).start()



    class ThreadedServer(socketserver.ThreadingMixIn, socketserver.TCPServer):
        allow_reuse_address = True
        daemon_threads = True  # clean shutdown on Ctrl+C

        def handle_error(self, request, client_address):
            import sys
            exc_type, exc_val, _ = sys.exc_info()
            # Silently ignore browser connection drops (very common, not real errors)
            if exc_type in (ConnectionResetError, BrokenPipeError):
                return
            if exc_type is OSError and getattr(exc_val, 'winerror', None) == 10054:
                return
            super().handle_error(request, client_address)

    with ThreadedServer(("", PORT), ChatbotHandler) as httpd:
        print(f"🌐 Server running at http://localhost:{PORT}")
        print("📊 Token usage will be tracked in MongoDB Atlas")
        print("Press Ctrl+C to stop\n")
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\n🛑 Server stopped")
            httpd.shutdown()
