"""
File handler service - Upload, validate, parse, store, fetch, and scraper utilities
"""
import os
import re
import shutil
import tempfile
import urllib.request
import urllib.error
from pathlib import Path
from datetime import datetime, timezone
from bs4 import BeautifulSoup
from botai.config.database import generate_id
from botai.config import settings
from botai.utils.validators import validate_file_type

class FileHandler:
    """Handles file uploads, validations, parsing, Google Drive fetching, and Web scraping"""
    
    @staticmethod
    def get_file_type(filename: str) -> str:
        """Determine file type by extension"""
        ext = filename.lower().split('.')[-1] if '.' in filename else ''
        for file_type, extensions in settings.ALLOWED_FILE_TYPES.items():
            if ext in extensions:
                return file_type
        return 'unknown'
    
    @staticmethod
    def validate_file(filename: str, file_size: int) -> tuple[bool, str]:
        """
        Validate file before upload
        Returns: (is_valid: bool, error_message: str)
        """
        if not filename:
            return False, "Filename is required"
        if file_size == 0:
            return False, "File is empty"
        if file_size > settings.MAX_FILE_SIZE_BYTES:
            max_mb = settings.MAX_FILE_SIZE_BYTES // (1024 * 1024)
            return False, f"File too large (max {max_mb}MB, got {file_size // (1024*1024)}MB)"
        
        file_type = FileHandler.get_file_type(filename)
        if file_type == 'unknown':
            ext = filename.split('.')[-1]
            return False, f"File type not allowed: .{ext}"
        
        return True, ""
    
    @staticmethod
    def save_file(file_data: bytes, filename: str, user_id: str, db) -> tuple[bool, str]:
        """
        Save uploaded file to disk and database
        Returns: (success: bool, file_id_or_error: str)
        """
        try:
            is_valid, error_msg = FileHandler.validate_file(filename, len(file_data))
            if not is_valid:
                return False, error_msg
            
            file_type = FileHandler.get_file_type(filename)
            dest_dir = settings.UPLOAD_DIR / file_type
            dest_dir.mkdir(parents=True, exist_ok=True)
            
            # Generate unique filename
            unique_filename = f"{generate_id()}_{filename}"
            dest_path = dest_dir / unique_filename
            
            # Save file to disk
            with open(dest_path, 'wb') as f:
                f.write(file_data)
            
            size_bytes = dest_path.stat().st_size
            
            # Save metadata to MySQL
            file_id = generate_id()
            file_doc = {
                '_id': file_id,
                'user_id': user_id,
                'filename': filename,
                'file_type': file_type,
                'size_bytes': size_bytes,
                'path': str(dest_path),
                'created_at': datetime.now(timezone.utc)
            }
            db.files.insert_one(file_doc)
            
            print(f"[FileHandler] File saved: {filename} ({file_type})")
            return True, str(file_id)
        except Exception as e:
            print(f"[FileHandler] Error uploading file: {e}")
            return False, f"Error uploading file: {str(e)}"
    
    @staticmethod
    def get_file_path(file_id: str, db) -> tuple[bool, str]:
        """
        Get file path from database
        Returns: (success: bool, path_or_error: str)
        """
        try:
            file_doc = db.files.find_one({'_id': file_id})
            if not file_doc:
                return False, f"File not found: {file_id}"
            return True, file_doc['path']
        except Exception as e:
            return False, f"Error retrieving file: {str(e)}"
    
    @staticmethod
    def get_file_content(file_id: str, db) -> tuple[bool, str]:
        """
        Read file content (supports PDF, Word, Excel, CSV, and text-based files)
        Returns: (success: bool, content_or_error: str)
        """
        try:
            success, path = FileHandler.get_file_path(file_id, db)
            if not success:
                return False, path
            
            content = FileHandler.parse_file(path)
            if content.startswith('[') and content.endswith(']'):
                # Indicates error returned by parser
                return False, content
            return True, content[:10000]
        except Exception as e:
            return False, f"Error reading file: {str(e)}"
    
    @staticmethod
    def delete_file(file_id: str, user_id: str, db) -> tuple[bool, str]:
        """
        Delete file and metadata
        Returns: (success: bool, message_or_error: str)
        """
        try:
            file_doc = db.files.find_one({
                '_id': file_id,
                'user_id': user_id
            })
            if not file_doc:
                return False, "File not found or access denied"

            # Delete physical file
            if os.path.exists(file_doc['path']):
                os.remove(file_doc['path'])
                print(f"[FileHandler] Deleted file: {file_doc['path']}")

            # Delete metadata
            db.files.delete_one({'_id': file_id})
            return True, "File deleted successfully"
        except Exception as e:
            return False, f"Error deleting file: {str(e)}"
    
    @staticmethod
    def list_user_files(user_id: str, db) -> list:
        """
        Get all files uploaded by a user
        Returns: List of file documents
        """
        try:
            files = db.files.find({
                'user_id': user_id
            }).sort('created_at', -1)
            files = list(files)
            return [
                {
                    'file_id': str(f['_id']),
                    'filename': f['filename'],
                    'file_type': f['file_type'],
                    'size_mb': round(f['size_bytes'] / (1024*1024), 2),
                    'created_at': f['created_at'].isoformat()
                }
                for f in files
            ]
        except Exception as e:
            print(f"[FileHandler] Error listing files: {e}")
            return []

    # ========== SCRAPING & TELEMETRY PARSING EXTENSIONS ==========

    @staticmethod
    def parse_file(filepath: str) -> str:
        """
        Extract readable plain text content from a file depending on its extension.
        Supports: PDF, DOCX, DOC, XLSX, XLS, and text files.
        """
        lower_path = filepath.lower()
        text = ""
        
        # PDF Parsing
        if lower_path.endswith('.pdf'):
            try:
                from pdfminer.high_level import extract_text as pdf_extract
                text = pdf_extract(filepath)
            except ImportError:
                text = "[PDF found but pdfminer.six not installed. Run: pip install pdfminer.six]"
            except Exception as e:
                text = f"[PDF extraction error: {e}]"
                
        # Word Doc Parsing
        elif lower_path.endswith('.docx') or lower_path.endswith('.doc'):
            try:
                import mammoth
                with open(filepath, 'rb') as f:
                    text = mammoth.extract_raw_text(f).value
            except ImportError:
                try:
                    from docx import Document
                    text = '\n'.join([p.text for p in Document(filepath).paragraphs])
                except ImportError:
                    text = "[Word file found but mammoth/python-docx not installed.]"
            except Exception as e:
                text = f"[Word extraction error: {e}]"
                
        # Excel XLSX Parsing
        elif lower_path.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, data_only=True)
                sheets_text = []
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    sheet_lines = [f"--- Sheet: {sheet_name} ---"]
                    for row in sheet.iter_rows(values_only=True):
                        if row and any(cell is not None and str(cell).strip() != "" for cell in row):
                            row_str = " | ".join(str(cell).strip() if cell is not None else "" for cell in row)
                            sheet_lines.append(row_str)
                    if len(sheet_lines) > 1:
                        sheets_text.append("\n".join(sheet_lines))
                text = "\n\n".join(sheets_text)
            except Exception as e:
                text = f"[Excel extraction error (.xlsx): {e}]"
                
        # Excel XLS Parsing
        elif lower_path.endswith('.xls'):
            try:
                import xlrd
                wb = xlrd.open_workbook(filepath)
                sheets_text = []
                for sheet_idx in range(wb.nsheets):
                    sheet = wb.sheet_by_index(sheet_idx)
                    sheet_lines = [f"--- Sheet: {sheet.name} ---"]
                    for row_idx in range(sheet.nrows):
                        row = sheet.row_values(row_idx)
                        if row and any(cell is not None and str(cell).strip() != "" for cell in row):
                            row_str = " | ".join(str(cell).strip() if cell is not None else "" for cell in row)
                            sheet_lines.append(row_str)
                    if len(sheet_lines) > 1:
                        sheets_text.append("\n".join(sheet_lines))
                text = "\n\n".join(sheets_text)
            except Exception as e:
                text = f"[Excel extraction error (.xls): {e}]"
                
        # Plain Text / Scripts Parsing
        elif any(lower_path.endswith(ext) for ext in ['.txt', '.csv', '.md', '.py', '.js', '.json', '.xml', '.html', '.log']):
            try:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    text = f.read()
            except Exception as e:
                text = f"[Text read error: {e}]"
        else:
            try:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    raw = f.read()
                text = raw if raw.strip() else ""
            except Exception:
                text = ""
        return text

    @staticmethod
    def extract_gdrive_file_id(url: str) -> str:
        """Extract Google Drive file/folder ID from multiple URL patterns"""
        patterns = [
            r'/file/d/([a-zA-Z0-9_-]{10,})',
            r'id=([a-zA-Z0-9_-]{10,})',
            r'/d/([a-zA-Z0-9_-]{10,})',
            r'/folders/([a-zA-Z0-9_-]{10,})',
        ]
        for pattern in patterns:
            m = re.search(pattern, url)
            if m:
                return m.group(1)
        return None

    @staticmethod
    def fetch_gdrive_text(url: str) -> str:
        """Download public Drive documents or folders — backward-compatible wrapper around fetch_gdrive_documents."""
        result = FileHandler.fetch_gdrive_documents(url)
        if result.get('error'):
            return f"[{result['error']}]"
        if not result.get('files'):
            return "[No readable documents found in the Google Drive link.]"
        # Flatten all files into a single text blob (for inline chat URL detection)
        parts = []
        for f in result['files']:
            if f.get('text'):
                parts.append(f"--- File: {f['name']} ---\n{f['text'].strip()}")
        combined = "\n\n".join(parts)
        return combined[:10000] if combined else "[No readable text found in the Google Drive link.]"

    @staticmethod
    def fetch_gdrive_documents(url: str, max_files: int = 30, per_file_chars: int = 15000, total_chars: int = 50000) -> dict:
        """
        Enhanced Google Drive document fetcher for the knowledge base feature.

        Returns a structured dict:
        {
          'files': [{'name': str, 'text': str, 'size_chars': int, 'type': str}],
          'total_chars': int,
          'files_loaded': int,
          'files_skipped': int,
          'error': str or None
        }

        Supports: PDF, DOCX, DOC, XLSX, XLS, TXT, CSV, MD, code files,
                  and Google Docs/Sheets/Slides (exported via Google's export API).
        """
        file_id = FileHandler.extract_gdrive_file_id(url)
        if not file_id:
            return {'files': [], 'total_chars': 0, 'files_loaded': 0, 'files_skipped': 0,
                    'error': 'Could not extract a file/folder ID from the provided Google Drive URL.'}

        try:
            import gdown
        except ImportError:
            return {'files': [], 'total_chars': 0, 'files_loaded': 0, 'files_skipped': 0,
                    'error': 'gdown library not installed on this server. Run: pip install gdown'}

        is_folder = '/folders/' in url
        is_gdoc   = 'docs.google.com/document' in url
        is_gsheet = 'docs.google.com/spreadsheets' in url
        is_gslide = 'docs.google.com/presentation' in url
        tmp_dir   = tempfile.mkdtemp()
        all_files = []
        files_skipped = 0

        try:
            # ── Google Docs / Sheets / Slides (native Google format) ──────────────
            if is_gdoc or is_gsheet or is_gslide:
                if is_gdoc:
                    export_url = f"https://docs.google.com/document/d/{file_id}/export?format=txt"
                    ext = '.txt'
                elif is_gsheet:
                    export_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=csv"
                    ext = '.csv'
                else:
                    export_url = f"https://docs.google.com/presentation/d/{file_id}/export?format=txt"
                    ext = '.txt'

                dest = os.path.join(tmp_dir, f'gdoc_export{ext}')
                print(f"[Drive KB] Exporting native Google Doc/Sheet/Slide: {export_url}")
                try:
                    req = urllib.request.Request(export_url, headers={'User-Agent': 'Mozilla/5.0'})
                    with urllib.request.urlopen(req, timeout=20) as resp:
                        with open(dest, 'wb') as f:
                            f.write(resp.read())
                    text = FileHandler.parse_file(dest)
                    name = url.split('/')[-2] if '/' in url else 'google_doc'
                    if text and text.strip():
                        all_files.append({'name': name + ext, 'text': text[:per_file_chars], 'size_chars': min(len(text), per_file_chars), 'type': 'gdoc'})
                    else:
                        all_files.append({'name': name + ext, 'text': '', 'size_chars': 0, 'type': 'gdoc'})
                except Exception as e:
                    return {'files': [], 'total_chars': 0, 'files_loaded': 0, 'files_skipped': 0,
                            'error': f'Failed to export Google Doc: {e}. Make sure sharing is set to "Anyone with the link".'}

            # ── Google Drive FOLDER ───────────────────────────────────────────────
            elif is_folder:
                folder_path = os.path.join(tmp_dir, 'drive_folder')
                os.makedirs(folder_path, exist_ok=True)
                print(f"[Drive KB] Downloading FOLDER ID: {file_id}")
                try:
                    gdown.download_folder(id=file_id, output=folder_path, quiet=True, use_cookies=False)
                except Exception as e:
                    shutil.rmtree(tmp_dir, ignore_errors=True)
                    return {'files': [], 'total_chars': 0, 'files_loaded': 0, 'files_skipped': 0,
                            'error': f'Google Drive folder download failed: {e}. Share the folder as "Anyone with the link".'}

                disk_files = []
                for root, dirs, files in os.walk(folder_path):
                    for fname in files:
                        disk_files.append(os.path.join(root, fname))

                if not disk_files:
                    shutil.rmtree(tmp_dir, ignore_errors=True)
                    return {'files': [], 'total_chars': 0, 'files_loaded': 0, 'files_skipped': 0,
                            'error': 'Google Drive folder appears empty or is restricted. Share it as "Anyone with the link".'}

                print(f"[Drive KB] Found {len(disk_files)} files in folder")
                for fpath in disk_files[:max_files]:
                    fname = os.path.basename(fpath)
                    # Skip hidden/system files
                    if fname.startswith('.') or fname.startswith('~'):
                        files_skipped += 1
                        continue
                    text = FileHandler.parse_file(fpath)
                    if text and text.strip():
                        all_files.append({
                            'name': fname,
                            'text': text[:per_file_chars],
                            'size_chars': min(len(text), per_file_chars),
                            'type': fname.rsplit('.', 1)[-1].lower() if '.' in fname else 'unknown'
                        })
                    else:
                        files_skipped += 1

                if len(disk_files) > max_files:
                    files_skipped += len(disk_files) - max_files

            # ── Google Drive SINGLE FILE ──────────────────────────────────────────
            else:
                dest_file = os.path.join(tmp_dir, 'drive_file')
                print(f"[Drive KB] Downloading FILE ID: {file_id}")
                try:
                    gdown.download(id=file_id, output=dest_file, quiet=True, use_cookies=False)
                except Exception as e:
                    # Fallback direct URL
                    direct_url = f"https://docs.google.com/uc?export=download&id={file_id}"
                    try:
                        req = urllib.request.Request(direct_url, headers={'User-Agent': 'Mozilla/5.0'})
                        with urllib.request.urlopen(req, timeout=20) as resp:
                            with open(dest_file, 'wb') as f:
                                f.write(resp.read())
                    except Exception as direct_err:
                        shutil.rmtree(tmp_dir, ignore_errors=True)
                        return {'files': [], 'total_chars': 0, 'files_loaded': 0, 'files_skipped': 0,
                                'error': f'File download failed: {direct_err}. Make sure it is shared as "Anyone with the link".'}

                if not os.path.exists(dest_file) or os.path.getsize(dest_file) == 0:
                    shutil.rmtree(tmp_dir, ignore_errors=True)
                    return {'files': [], 'total_chars': 0, 'files_loaded': 0, 'files_skipped': 0,
                            'error': 'File downloaded but appears empty or invalid.'}

                text = FileHandler.parse_file(dest_file)
                # Try to get original filename from URL
                fname = url.split('/')[-1].split('?')[0] or 'drive_file'
                if text and text.strip():
                    all_files.append({
                        'name': fname,
                        'text': text[:per_file_chars],
                        'size_chars': min(len(text), per_file_chars),
                        'type': fname.rsplit('.', 1)[-1].lower() if '.' in fname else 'unknown'
                    })
                else:
                    files_skipped += 1

        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

        # Apply total character budget across all files
        used_chars = 0
        trimmed_files = []
        for f in all_files:
            remaining = total_chars - used_chars
            if remaining <= 0:
                files_skipped += 1
                continue
            if f['size_chars'] > remaining:
                f['text'] = f['text'][:remaining]
                f['size_chars'] = remaining
            trimmed_files.append(f)
            used_chars += f['size_chars']

        return {
            'files': trimmed_files,
            'total_chars': used_chars,
            'files_loaded': len(trimmed_files),
            'files_skipped': files_skipped,
            'error': None
        }

    @staticmethod
    def fetch_url_text(url: str) -> str:
        """Scrapes web page URL, strips standard headers/scripts, and returns page text context"""
        try:
            url = url.strip()
            req = urllib.request.Request(
                url,
                headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                    'Accept-Language': 'en-US,en;q=0.5'
                }
            )
            with urllib.request.urlopen(req, timeout=8) as response:
                html = response.read()
                soup = BeautifulSoup(html, 'html.parser')
                
                # Decompose script/style blocks
                for element in soup(["script", "style", "noscript", "iframe", "header", "footer", "nav"]):
                    element.decompose()
                    
                text = soup.get_text()
                lines = (line.strip() for line in text.splitlines())
                chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
                clean_text = '\n'.join(chunk for chunk in chunks if chunk)
                return clean_text[:6000]
        except Exception as e:
            print(f"[Web Fetcher] Error fetching {url}: {e}")
            return f"[Failed to retrieve webpage content from {url} due to error: {e}]"
